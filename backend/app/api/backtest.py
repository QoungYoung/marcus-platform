# -*- coding: utf-8 -*-
"""
Backtest API - AI 交易回测系统
支持创建/列表/进度/结果/取消 + SSE 实时流式进度
"""
import asyncio
import json
import logging
import re
import uuid
from datetime import date, datetime
from typing import Optional, Dict, List

import pandas as pd
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from app.database import SessionLocal
from app.models.backtest_orm import (
    BacktestTask, BacktestDailyLog, BacktestTrade,
    BacktestPosition, BacktestEquitySnapshot, BacktestMonthlyMetric,
)
from app.services.backtest_engine import backtest_engine

router = APIRouter(prefix="/backtest", tags=["Backtest"])

logger = logging.getLogger(__name__)


# ── Pydantic Schemas ──

class CreateBacktestRequest(BaseModel):
    name: str = Field(..., description="任务名称")
    start_date: str = Field(..., pattern=r"^\d{4}-\d{2}-\d{2}$", description="起始日期 YYYY-MM-DD")
    end_date: str = Field(..., pattern=r"^\d{4}-\d{2}-\d{2}$", description="结束日期 YYYY-MM-DD")
    initial_capital: float = Field(1_000_000, ge=10000, description="初始资金")
    include_chinext: bool = Field(False, description="是否包含创业板股票(300/301开头)")
    model: Optional[str] = Field("deepseek-v4-pro", description="AI模型: deepseek-v4-pro / deepseek-v4-flash")
    thinking_level: Optional[str] = Field("high", description="思考等级: high / medium / low")


class BacktestTaskResponse(BaseModel):
    id: str
    name: str
    start_date: str
    end_date: str
    initial_capital: float
    status: str
    current_day: Optional[str] = None
    total_days: int
    completed_days: int
    progress: float
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    error_message: Optional[str] = None
    created_at: str

    class Config:
        from_attributes = True


class BacktestTaskListResponse(BaseModel):
    tasks: list
    total: int


# ── Helpers ──

def _task_to_dict(task: BacktestTask) -> dict:
    return {
        "id": task.id,
        "name": task.name,
        "start_date": task.start_date.isoformat() if task.start_date else "",
        "end_date": task.end_date.isoformat() if task.end_date else "",
        "initial_capital": task.initial_capital,
        "status": task.status,
        "current_day": task.current_day.isoformat() if task.current_day else None,
        "total_days": task.total_days,
        "completed_days": task.completed_days,
        "progress": task.progress,
        "started_at": task.started_at.isoformat() if task.started_at else None,
        "completed_at": task.completed_at.isoformat() if task.completed_at else None,
        "error_message": task.error_message,
        "created_at": task.created_at.isoformat() if task.created_at else "",
        "model": getattr(task, "model_name", "deepseek-v4-pro") or "deepseek-v4-pro",
        "thinking_level": getattr(task, "thinking_level", "high") or "high",
    }


# ── API Endpoints ──

@router.post("/create")
async def create_backtest(req: CreateBacktestRequest):
    """创建回测任务"""
    task_id = str(uuid.uuid4())[:12]

    db = SessionLocal()
    try:
        task = BacktestTask(
            id=task_id,
            name=req.name,
            start_date=date.fromisoformat(req.start_date),
            end_date=date.fromisoformat(req.end_date),
            initial_capital=req.initial_capital,
            include_chinext=req.include_chinext,
            model_name=req.model or "deepseek-v4-pro",
            thinking_level=req.thinking_level or "high",
            status="pending",
        )
        db.add(task)
        db.commit()
        return {"success": True, "task_id": task_id, "message": "回测任务已创建"}
    finally:
        db.close()


@router.post("/{task_id}/start")
async def start_backtest(task_id: str):
    """验证任务状态（实际启动由 stream 端点完成）"""
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")
        if task.status not in ("pending", "failed", "cancelled"):
            raise HTTPException(400, f"任务状态为 {task.status}，无法启动")
    finally:
        db.close()
    return {"success": True, "message": "回测就绪，请连接 stream", "task_id": task_id}


# 全局：存每个任务的 event 队列，解决 start/stream 竞态
_stream_queues: Dict[str, asyncio.Queue] = {}


@router.get("/{task_id}/stream")
async def stream_backtest(task_id: str):
    """SSE 流式获取回测进度（统一入口，这里才真正启动回测）"""
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")
    finally:
        db.close()

    queue = asyncio.Queue()
    # 替换旧队列（如果有，说明重连了）
    old_q = _stream_queues.get(task_id)
    if old_q:
        try:
            old_q.put_nowait(None)  # 通知旧连接关闭
        except Exception:
            pass
    _stream_queues[task_id] = queue

    # 如果任务已完成，直接推送结果
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if task and task.status in ("completed", "failed", "cancelled"):
            queue.put_nowait({
                "event": "complete" if task.status == "completed" else "error",
                "message": f"任务已{task.status}",
                "progress": task.progress,
                "data": {"status": task.status},
            })
            queue.put_nowait(None)
    finally:
        db.close()

    # 如果未运行也未完成，启动回测（引擎内部推全局队列）
    if task and task.status not in ("completed", "failed", "cancelled", "running"):
        asyncio.create_task(_run_backtest_async(task_id))

    async def event_generator():
        try:
            while True:
                event = await queue.get()
                if event is None:
                    break
                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
            yield f"data: {json.dumps({'event': 'done', 'message': '', 'progress': 100, 'data': {}})}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            _stream_queues.pop(task_id, None)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@router.post("/{task_id}/cancel")
async def cancel_backtest(task_id: str):
    """取消回测任务"""
    backtest_engine.cancel_task(task_id)

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if task and task.status == "running":
            task.status = "cancelled"
            task.completed_at = datetime.now()
            db.commit()
    finally:
        db.close()

    return {"success": True, "message": "回测已取消"}


@router.get("/tasks")
async def list_backtest_tasks(
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
):
    """获取回测任务列表"""
    db = SessionLocal()
    try:
        total = db.query(BacktestTask).count()
        tasks = (
            db.query(BacktestTask)
            .order_by(BacktestTask.created_at.desc())
            .offset(offset)
            .limit(limit)
            .all()
        )
        return {
            "tasks": [_task_to_dict(t) for t in tasks],
            "total": total,
        }
    finally:
        db.close()


@router.get("/{task_id}")
async def get_backtest_detail(task_id: str):
    """获取回测任务详情（含指标摘要）"""
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        result = _task_to_dict(task)

        # 获取权益曲线
        equity = (
            db.query(BacktestEquitySnapshot)
            .filter(BacktestEquitySnapshot.task_id == task_id)
            .order_by(BacktestEquitySnapshot.trade_date)
            .all()
        )
        result["equity_curve"] = [
            {
                "date": e.trade_date.isoformat(),
                "total_asset": e.total_asset,
                "cost_based_asset": e.cost_based_asset or 0,
                "available_cash": e.available_cash,
                "position_value": e.position_value,
                "cost_value": e.cost_value or 0,
                "float_pnl": e.float_pnl or 0,
                "daily_pct": e.daily_pct or 0,
                "cost_based_return": e.cost_based_return or 0,
                "daily_return": e.daily_return,
                "cumulative_return": e.cumulative_return,
                "baseline_return": e.baseline_return or 0,
            }
            for e in equity
        ]

        # 获取交易记录
        trades = (
            db.query(BacktestTrade)
            .filter(BacktestTrade.task_id == task_id)
            .order_by(BacktestTrade.trade_date, BacktestTrade.id)
            .all()
        )
        result["trades"] = [
            {
                "id": t.id,
                "trade_date": t.trade_date.isoformat(),
                "symbol": t.symbol,
                "stock_name": t.stock_name or "",
                "direction": t.direction,
                "price": t.price,
                "avg_cost": round(
                    (t.price - (t.profit or 0) / t.volume) if t.direction == 'sell' and t.volume > 0 else t.price, 2
                ),
                "volume": t.volume,
                "amount": t.amount,
                "commission": t.commission or 0,
                "profit": t.profit or 0,
                "profit_pct": t.profit_pct or 0,
                "reason": t.reason,
                # 滑点 + 税费明细 (新增)
                "phase_time": t.phase_time or "",
                "signal_price": round(float(t.signal_price or 0), 2),
                "actual_price": round(float(t.actual_price or 0) or t.price, 2),
                "slippage_pct": round(float(t.slippage_pct or 0), 4),
                "stamp_tax": round(float(t.stamp_tax or 0), 2),
                "transfer_fee": round(float(t.transfer_fee or 0), 2),
                "net_profit": round(float(t.net_profit or 0), 2),
            }
            for t in trades
        ]

        # 获取月度指标
        monthly = (
            db.query(BacktestMonthlyMetric)
            .filter(BacktestMonthlyMetric.task_id == task_id)
            .order_by(BacktestMonthlyMetric.month)
            .all()
        )
        if not monthly:
            monthly = _compute_monthly_metrics_from_db(db, task_id)
        result["monthly_metrics"] = [
            {
                "month": m.month,
                "return_pct": m.return_pct,
                "trades_count": m.trades_count,
                "win_count": m.win_count,
                "win_rate": m.win_rate,
                "max_drawdown": m.max_drawdown,
            }
            for m in monthly
        ]

        # 计算汇总指标
        if equity:
            initial = task.initial_capital
            final = equity[-1].total_asset
            total_return = (final / initial - 1) * 100
            years = len(equity) / 252 if len(equity) > 0 else 1
            annual_return = ((1 + total_return / 100) ** (1 / years) - 1) * 100 if years > 0 else 0

            # 最大回撤
            peak = 0.0
            max_dd = 0.0
            cum = 1.0
            for e in equity:
                cum = 1 + e.cumulative_return / 100
                if cum > peak:
                    peak = cum
                dd = (peak - cum) / peak * 100
                if dd > max_dd:
                    max_dd = dd

            buy_trades = [t for t in trades if t.direction == "buy"]
            sell_trades = [t for t in trades if t.direction == "sell"]
            win_count = len(sell_trades)  # simplified - all sells counted

            result["metrics"] = {
                "total_return": round(total_return, 2),
                "annual_return": round(annual_return, 2),
                "max_drawdown": round(max_dd, 2),
                "total_trades": len(trades),
                "buy_count": len(buy_trades),
                "sell_count": len(sell_trades),
                "final_equity": round(final, 2),
                "total_days": task.total_days,
            }

        # 获取每日日志
        logs = (
            db.query(BacktestDailyLog)
            .filter(BacktestDailyLog.task_id == task_id)
            .order_by(BacktestDailyLog.trade_date, BacktestDailyLog.id)
            .limit(200)
            .all()
        )
        result["daily_logs"] = [
            {
                "id": l.id,
                "trade_date": l.trade_date.isoformat(),
                "day_index": l.day_index,
                "phase": l.phase,
                "phase_time": l.phase_time,
                "event_type": l.event_type,
                "content": l.content[:300] if l.content else "",
                "has_prompt_snapshot": bool(l.metadata_json and l.metadata_json.get("prompt_snapshot")),
                "has_full_reply": bool(l.metadata_json and l.metadata_json.get("full_reply")),
            }
            for l in logs
        ]

        # 获取最后持仓
        positions = (
            db.query(BacktestPosition)
            .filter(BacktestPosition.task_id == task_id)
            .order_by(BacktestPosition.trade_date.desc())
            .limit(50)
            .all()
        )
        # 取最后一天的持仓
        last_date_positions = {}
        last_date = None
        if positions:
            last_date = positions[0].trade_date
            for p in positions:
                if p.trade_date == last_date:
                    last_date_positions[p.symbol] = p

        # 补全名称/T+1 状态
        from app.services.backtest_engine import backtest_engine
        from app.services.local_data_provider import local_data
        engine = backtest_engine._engines.get(task_id)
        name_cache: dict = {}
        result_positions = []
        for sym, p in last_date_positions.items():
            # 名称
            if sym not in name_cache:
                try:
                    name_cache[sym] = local_data.get_stock_name(sym) or ""
                except Exception:
                    name_cache[sym] = ""
            # T+1 状态（与 /sandbox/positions 字段对齐）
            t1 = {"locked": False, "last_buy_date": None, "unlock_date": None, "reason": "引擎已结束"}
            if engine:
                try:
                    engine_sym = sym if sym.startswith(("SH", "SZ")) else (
                        "SH" + sym.split(".")[0] if sym.endswith(".SH")
                        else "SZ" + sym.split(".")[0] if sym.endswith(".SZ")
                        else sym
                    )
                    t1 = engine.get_t1_status(engine_sym)
                except Exception:
                    pass
            # 建仓日（从该 symbol 的所有 buy 记录取最早 trade_date）
            first_buy = (
                db.query(BacktestTrade)
                .filter(BacktestTrade.task_id == task_id,
                        BacktestTrade.symbol == sym,
                        BacktestTrade.direction == "buy")
                .order_by(BacktestTrade.trade_date.asc())
                .first()
            )
            entry_date = first_buy.trade_date.isoformat() if first_buy else None
            # 持仓天数（自然日）
            holding_days = 0
            if entry_date and last_date:
                from datetime import date as _date
                holding_days = (last_date - _date.fromisoformat(entry_date)).days

            result_positions.append({
                "symbol": p.symbol,
                "stock_name": name_cache[sym],
                "trade_date": p.trade_date.isoformat(),
                "volume": p.volume,
                "avg_cost": p.avg_cost,
                "current_price": p.current_price,
                "cost_value": round(p.avg_cost * p.volume, 2),
                "market_value": p.market_value,
                "float_pnl": p.float_pnl,
                "float_pnl_pct": p.float_pnl_pct,
                "entry_date": entry_date,
                "holding_days": holding_days,
                "t1_status": t1,
            })
        result["final_positions"] = result_positions

        return result
    finally:
        db.close()


@router.get("/{task_id}/trades")
async def get_backtest_trades(
    task_id: str,
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=200, description="每页条数"),
    direction: Optional[str] = Query(None, description="方向: buy / sell"),
    keyword: Optional[str] = Query(None, description="搜索关键词 (代码/名称/理由)"),
    start_date: Optional[str] = Query(None, description="起始日期 YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="结束日期 YYYY-MM-DD"),
):
    """获取回测交易明细（分页 + 条件搜索）"""
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        from sqlalchemy import func, or_

        # 构建过滤条件
        filters = [BacktestTrade.task_id == task_id]

        if direction:
            filters.append(BacktestTrade.direction == direction)

        if keyword:
            kw_pattern = f"%{keyword}%"
            filters.append(or_(
                BacktestTrade.symbol.ilike(kw_pattern),
                BacktestTrade.stock_name.ilike(kw_pattern),
                BacktestTrade.reason.ilike(kw_pattern),
            ))

        if start_date:
            try:
                filters.append(BacktestTrade.trade_date >= date.fromisoformat(start_date))
            except ValueError:
                raise HTTPException(400, f"无效的起始日期: {start_date}")

        if end_date:
            try:
                filters.append(BacktestTrade.trade_date <= date.fromisoformat(end_date))
            except ValueError:
                raise HTTPException(400, f"无效的结束日期: {end_date}")

        # 总数
        total = db.query(func.count(BacktestTrade.id)).filter(*filters).scalar() or 0

        # 分页查询
        offset = (page - 1) * page_size
        trades = (
            db.query(BacktestTrade)
            .filter(*filters)
            .order_by(BacktestTrade.trade_date.desc(), BacktestTrade.id.desc())
            .offset(offset)
            .limit(page_size)
            .all()
        )

        result_trades = [
            {
                "id": t.id,
                "trade_date": t.trade_date.isoformat(),
                "symbol": t.symbol,
                "stock_name": t.stock_name or "",
                "direction": t.direction,
                "price": t.price,
                "avg_cost": round(
                    (t.price - (t.profit or 0) / t.volume) if t.direction == 'sell' and t.volume > 0 else t.price, 2
                ),
                "volume": t.volume,
                "amount": t.amount,
                "commission": t.commission or 0,
                "profit": t.profit or 0,
                "profit_pct": t.profit_pct or 0,
                "reason": t.reason,
                "phase_time": t.phase_time or "",
                "signal_price": round(float(t.signal_price or 0), 2),
                "actual_price": round(float(t.actual_price or 0) or t.price, 2),
                "slippage_pct": round(float(t.slippage_pct or 0), 4),
                "stamp_tax": round(float(t.stamp_tax or 0), 2),
                "transfer_fee": round(float(t.transfer_fee or 0), 2),
                "net_profit": round(float(t.net_profit or 0), 2),
            }
            for t in trades
        ]

        return {
            "trades": result_trades,
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": max(1, (total + page_size - 1) // page_size),
        }
    finally:
        db.close()


@router.get("/{task_id}/equity-csv")
async def export_equity_csv(task_id: str):
    """导出回测实时权益曲线为 CSV 文件"""
    from io import StringIO
    import csv

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        equity = (
            db.query(BacktestEquitySnapshot)
            .filter(BacktestEquitySnapshot.task_id == task_id)
            .order_by(BacktestEquitySnapshot.trade_date)
            .all()
        )
        if not equity:
            raise HTTPException(404, "该任务暂无权益数据")

        buf = StringIO()
        writer = csv.writer(buf)
        # 表头
        writer.writerow([
            "date", "total_asset", "available_cash", "position_value",
            "cost_value", "float_pnl", "cost_based_asset",
            "daily_pct", "daily_return", "cumulative_return",
            "cost_based_return", "baseline_return",
        ])
        # 数据行
        for e in equity:
            writer.writerow([
                e.trade_date.isoformat(),
                round(e.total_asset or 0, 2),
                round(e.available_cash or 0, 2),
                round(e.position_value or 0, 2),
                round(e.cost_value or 0, 2),
                round(e.float_pnl or 0, 2),
                round(e.cost_based_asset or 0, 2),
                round(e.daily_pct or 0, 4),
                round(e.daily_return or 0, 4),
                round(e.cumulative_return or 0, 4),
                round(e.cost_based_return or 0, 4),
                round(e.baseline_return or 0, 4),
            ])

        # 处理中文文件名 (RFC 5987)
        safe_name = task.name or f"backtest_{task_id[:8]}"
        ascii_name = f"equity_{task_id[:8]}.csv"
        utf8_name = f"equity_{safe_name}_{task_id[:8]}.csv"
        from urllib.parse import quote
        content_disposition = (
            f"attachment; filename=\"{ascii_name}\"; "
            f"filename*=UTF-8''{quote(utf8_name)}"
        )

        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": content_disposition,
                "Cache-Control": "no-cache",
            },
        )
    finally:
        db.close()




# ─────────────────────────────────────────────────────────────
# 月度指标 On-Demand 计算 (已取消任务缺少 BacktestMonthlyMetric 时用)
# ─────────────────────────────────────────────────────────────

class _MonthlyRow:
    """轻量月度指标行, 兼容 BacktestMonthlyMetric 的属性访问"""
    def __init__(self, month, return_pct, trades_count, win_count, win_rate, max_drawdown):
        self.month = month
        self.return_pct = return_pct
        self.trades_count = trades_count
        self.win_count = win_count
        self.win_rate = win_rate
        self.max_drawdown = max_drawdown


def _compute_monthly_metrics_from_db(db, task_id: str) -> List[_MonthlyRow]:
    """从 equity snapshots + trades 实时计算月度绩效"""
    from collections import defaultdict
    from sqlalchemy import func

    results: List[_MonthlyRow] = []

    # 1) 每月首/末资产 + 日收益
    equity = (
        db.query(BacktestEquitySnapshot)
        .filter(BacktestEquitySnapshot.task_id == task_id)
        .order_by(BacktestEquitySnapshot.trade_date)
        .all()
    )
    if not equity:
        return results

    month_first: dict = {}
    month_last: dict = {}
    month_daily: dict = defaultdict(list)

    for e in equity:
        m = e.trade_date.isoformat()[:7]
        if m not in month_first:
            month_first[m] = float(e.total_asset or 0)
        month_last[m] = float(e.total_asset or 0)
        month_daily[m].append(float(e.daily_pct or 0))

    # 2) 每月交易笔数 + 胜率
    trades_by_month: dict = defaultdict(lambda: {"total": 0, "wins": 0, "losses": 0})
    all_trades = (
        db.query(BacktestTrade)
        .filter(BacktestTrade.task_id == task_id)
        .all()
    )
    for t in all_trades:
        m = t.trade_date.isoformat()[:7]
        td = trades_by_month[m]
        if t.direction == "sell":
            td["total"] += 1
            p = float(t.profit or 0)
            if p > 0:
                td["wins"] += 1
            elif p < 0:
                td["losses"] += 1

    for month in sorted(month_first.keys()):
        start = month_first[month]
        end = month_last[month]
        month_return = (end / start - 1) * 100 if start > 0 else 0

        td = trades_by_month.get(month, {"total": 0, "wins": 0, "losses": 0})
        trades_cnt = td["total"]
        wins = td["wins"]
        losses = td["losses"]
        denom = wins + losses
        win_rate = round(wins / denom * 100, 2) if denom > 0 else 0

        # 月内最大回撤 (累乘日收益率)
        peak = 1.0
        max_dd = 0.0
        cum = 1.0
        for daily_pct in month_daily[month]:
            cum *= (1 + daily_pct / 100)
            if cum > peak:
                peak = cum
            dd = (peak - cum) / peak * 100 if peak > 0 else 0
            if dd > max_dd:
                max_dd = dd

        results.append(_MonthlyRow(
            month=month,
            return_pct=round(month_return, 4),
            trades_count=trades_cnt,
            win_count=wins,
            win_rate=win_rate,
            max_drawdown=round(max_dd, 4),
        ))

    return results


# ─────────────────────────────────────────────────────────────
# 全部导出 (Excel 多 Sheet, 一次拿到 4 个 CSV + 策略报告 + 汇总)
# ─────────────────────────────────────────────────────────────

@router.get("/{task_id}/export-all")
async def export_all_xlsx(task_id: str):
    """一键全量导出: 4 个数据 Sheet + 1 个策略报告 Sheet + 1 个汇总 Sheet
    适合一次性给分析师/合作方发完整数据包,无需逐个点

    ⚠️ 性能: openpyxl 是 CPU 密集型同步库, 写大文件会阻塞几秒。
    用 asyncio.to_thread 放到线程池,不阻塞 event loop,其他接口(SSE/查询)不被卡。
    """
    import asyncio
    from io import BytesIO
    from urllib.parse import quote

    def _build_xlsx():
        """同步函数,在线程池运行,避免阻塞 FastAPI event loop"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        db2 = SessionLocal()
        try:
            return _build_xlsx_sync(task_id, db2)
        finally:
            db2.close()

    # 在线程池执行 (FastAPI sync_workers 默认 40 个,够用)
    buf_content, filename, data_source = await asyncio.to_thread(_build_xlsx)

    content_disposition = (
        f"attachment; filename=\"{filename.split('_')[-1]}\"; "
        f"filename*=UTF-8''{quote(filename)}"
    )
    return StreamingResponse(
        iter([buf_content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": content_disposition,
            "Cache-Control": "no-cache",
            "X-Data-Source": data_source or "n/a",
            "Content-Length": str(len(buf_content)),
        },
    )


def _build_xlsx_sync(task_id: str, db):
    """在线程池内执行的 Excel 构建逻辑(同步)
    拆出来便于 asyncio.to_thread 调用
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        wb = Workbook()
        # 删除默认 Sheet
        wb.remove(wb.active)

        # ── 样式 ──
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="2C3E50")
        title_font = Font(bold=True, size=14, color="2C3E50")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _auto_width(ws, min_w=10, max_w=40):
            """根据内容自适应列宽 (中文字符按 2 算)"""
            for col_idx, col_cells in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = get_column_letter(col_idx)
                for cell in col_cells:
                    if cell.value is None:
                        continue
                    s = str(cell.value)
                    # 中文字符宽度近似
                    w = sum(2 if ord(c) > 127 else 1 for c in s)
                    if w > max_len:
                        max_len = w
                ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))

        def _write_header(ws, headers):
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = center
                c.border = border
            ws.row_dimensions[1].height = 24
            ws.freeze_panes = "A2"

        # ── Sheet 1: 任务汇总 ──
        ws = wb.create_sheet("任务汇总")
        rows = [
            ("任务ID", task.id),
            ("任务名称", task.name),
            ("回测区间", f"{task.start_date} ~ {task.end_date}"),
            ("初始资金", task.initial_capital),
            ("状态", task.status),
            ("包含创业板", "是" if getattr(task, "include_chinext", False) else "否"),
            ("交易日总数", task.total_days),
            ("已完成天数", task.completed_days),
            ("当前进度", f"{task.progress:.1f}%"),
            ("开始时间", task.started_at.isoformat() if task.started_at else ""),
            ("完成时间", task.completed_at.isoformat() if task.completed_at else ""),
            ("错误信息", task.error_message or ""),
        ]
        ws.cell(row=1, column=1, value="字段").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value="值").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        for ri, (k, v) in enumerate(rows, 2):
            ws.cell(row=ri, column=1, value=k).font = Font(bold=True)
            ws.cell(row=ri, column=2, value=v)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 50
        ws.freeze_panes = "A2"

        # ── Sheet 2: 逐笔交易明细 ──
        ws = wb.create_sheet("逐笔交易")
        trade_headers = [
            "交易日期", "信号时分", "合成时间",
            "股票代码", "股票名称", "方向",
            "信号价", "实际价", "滑点%",
            "数量", "金额",
            "手续费", "印花税", "过户费", "费用合计", "费用占比%",
            "毛盈亏", "盈亏%", "净盈亏",
            "T+0违规", "T+0说明",
            "交易理由",
        ]
        _write_header(ws, trade_headers)
        trades = (db.query(BacktestTrade)
                  .filter(BacktestTrade.task_id == task_id)
                  .order_by(BacktestTrade.trade_date, BacktestTrade.created_at)
                  .all())
        for ri, t in enumerate(trades, 2):
            td = t.trade_date.isoformat() if t.trade_date else ""
            pt = t.phase_time or ""
            dt = f"{td} {pt}:00" if pt else td
            amount = float(t.amount or 0)
            comm = float(t.commission or 0)
            stamp = float(t.stamp_tax or 0)
            trans = float(t.transfer_fee or 0)
            fee_total = comm + trans  # commission 字段已含印花税,这里不重复加
            fee_ratio = round(fee_total / amount * 100, 4) if amount > 0 else 0
            is_t0 = "✓" if getattr(t, "is_t0_violation", False) else ""
            row = [
                td, pt, dt,
                t.symbol, t.stock_name or "", t.direction,
                round(float(t.signal_price or 0), 3),
                round(float(t.actual_price or 0), 3),
                round(float(t.slippage_pct or 0), 4),
                int(t.volume or 0),
                round(amount, 2),
                round(comm, 2), round(stamp, 2), round(trans, 2),
                round(fee_total, 2), fee_ratio,
                round(float(t.profit or 0), 2),
                round(float(t.profit_pct or 0), 4),
                round(float(t.net_profit or 0), 2),
                is_t0,
                (t.t0_violation_note or "").replace("\n", " "),
                (t.reason or "").replace("\n", " "),
            ]
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.alignment = Alignment(vertical="center", wrap_text=(ci in (19, 21)))
        _auto_width(ws)

        # ── Sheet 3: 每日持仓快照 ──
        ws = wb.create_sheet("每日持仓")
        pos_headers = [
            "日期", "股票代码", "股票名称", "持仓数量", "平均成本", "收盘价",
            "市值", "浮盈", "浮盈%", "总资产", "仓位占比%",
        ]
        _write_header(ws, pos_headers)
        equity = {
            e.trade_date: e
            for e in db.query(BacktestEquitySnapshot)
            .filter(BacktestEquitySnapshot.task_id == task_id).all()
        }
        try:
            from app.services.local_data_provider import local_data
        except Exception:
            local_data = None
        positions = (db.query(BacktestPosition)
                     .filter(BacktestPosition.task_id == task_id)
                     .order_by(BacktestPosition.trade_date, BacktestPosition.symbol)
                     .all())
        for ri, p in enumerate(positions, 2):
            eq = equity.get(p.trade_date)
            total_asset = float(eq.total_asset) if eq and eq.total_asset else 0
            ratio = round(p.market_value / total_asset * 100, 4) \
                if total_asset > 0 and p.market_value else 0
            name = ""
            if local_data:
                try:
                    name = local_data.get_stock_name(p.symbol) or ""
                except Exception:
                    pass
            row = [
                p.trade_date.isoformat() if p.trade_date else "",
                p.symbol, name,
                int(p.volume or 0),
                round(float(p.avg_cost or 0), 3),
                round(float(p.current_price or 0), 3),
                round(float(p.market_value or 0), 2),
                round(float(p.float_pnl or 0), 2),
                round(float(p.float_pnl_pct or 0), 4),
                round(total_asset, 2),
                ratio,
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        _auto_width(ws)

        # ── Sheet 4: 每日权益曲线 ──
        ws = wb.create_sheet("权益曲线")
        eq_headers = [
            "日期", "总资产", "可用现金", "持仓市值",
            "持仓成本", "浮盈", "成本基准资产",
            "当日%", "累计%", "成本基准累计%", "资产指数",
        ]
        _write_header(ws, eq_headers)
        equities = (db.query(BacktestEquitySnapshot)
                    .filter(BacktestEquitySnapshot.task_id == task_id)
                    .order_by(BacktestEquitySnapshot.trade_date)
                    .all())
        for ri, e in enumerate(equities, 2):
            row = [
                e.trade_date.isoformat() if e.trade_date else "",
                round(float(e.total_asset or 0), 2),
                round(float(e.available_cash or 0), 2),
                round(float(e.position_value or 0), 2),
                round(float(e.cost_value or 0), 2),
                round(float(e.float_pnl or 0), 2),
                round(float(e.cost_based_asset or 0), 2),
                round(float(e.daily_pct or 0), 4),
                round(float(e.daily_return or 0), 4),
                round(float(e.cost_based_return or 0), 4),
                round(float(e.baseline_return or 0), 4),
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        _auto_width(ws)

        # ── Sheet 5: 月度绩效 ──
        ws = wb.create_sheet("月度绩效")
        mon_headers = ["月份", "月度收益%", "交易笔数", "盈利笔数", "胜率%", "月内最大回撤%"]
        _write_header(ws, mon_headers)
        metrics = (db.query(BacktestMonthlyMetric)
                   .filter(BacktestMonthlyMetric.task_id == task_id)
                   .order_by(BacktestMonthlyMetric.month)
                   .all())
        if not metrics:
            # 已取消的任务缺少月度指标 → 从 equity + trades 实时计算
            metrics = _compute_monthly_metrics_from_db(db, task_id)
        for ri, m in enumerate(metrics, 2):
            row = [
                m.month, round(float(m.return_pct or 0), 4),
                int(m.trades_count or 0), int(m.win_count or 0),
                round(float(m.win_rate or 0), 2),
                round(float(m.max_drawdown or 0), 4),
            ]
            for ci, val in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=val)
        _auto_width(ws)

        # ── Sheet 6: 基准指数 (尽力而为) ──
        ws = wb.create_sheet("基准指数")
        idx_headers = [
            "日期",
            "上证收盘", "上证%",
            "深证收盘", "深证%",
            "沪深300收盘", "沪深300%",
            "创业板收盘", "创业板%",
            "科创50收盘", "科创50%",
            "涨幅TOP1板块", "TOP1%",
            "涨幅TOP2板块", "TOP2%",
            "涨幅TOP3板块", "TOP3%",
            "跌幅TOP1板块", "BOT1%",
        ]
        _write_header(ws, idx_headers)
        # 复用 index-csv 端点的逻辑
        from pathlib import Path
        target_codes = ["000001.SH", "399001.SZ", "000300.SH", "399006.SZ", "000688.SH"]
        rows_by_date: Dict[str, Dict[str, dict]] = {}
        sector_data: Dict[str, Dict[str, float]] = {}
        data_source = ""
        # 方案 A: BacktestDailyLog
        try:
            logs = (db.query(BacktestDailyLog)
                    .filter(BacktestDailyLog.task_id == task_id)
                    .filter(BacktestDailyLog.event_type.in_(["scan_report", "pi_trade"]))
                    .filter(BacktestDailyLog.metadata_json.isnot(None))
                    .all())
            for log in logs:
                meta = log.metadata_json or {}
                if not meta.get("indices"):
                    continue
                td = log.trade_date.isoformat() if log.trade_date else ""
                if not td or td not in rows_by_date:
                    rows_by_date[td] = {}
                for idx in meta.get("indices", []):
                    code = idx.get("ts_code") or idx.get("code")
                    if code in target_codes:
                        rows_by_date[td][code] = {
                            "close": float(idx.get("close") or idx.get("current_price") or 0),
                            "pct_chg": float(idx.get("change_pct") or idx.get("pct_chg") or 0),
                        }
            if rows_by_date:
                data_source = "backtest_daily_log_cache"
        except Exception:
            pass
        # 方案 B: index_1min 推算
        if not rows_by_date:
            try:
                base = Path("f:/pythonProject/AITrade/marcus-platform/data/backtest/指数数据/index_1min")
                from datetime import date as dt_date, timedelta
                start = task.start_date or dt_date(2026, 1, 1)
                end = task.end_date or dt_date.today()
                cur = start
                while cur <= end:
                    td_str = cur.isoformat()
                    rows_by_date[td_str] = {}
                    for code in target_codes:
                        fp = base / f"{code}.parquet"
                        if not fp.exists():
                            continue
                        try:
                            df = pd.read_parquet(fp)
                            day_df = df.loc[td_str] if td_str in df.index.get_level_values(0) else None
                            if day_df is not None and not day_df.empty:
                                last = day_df.iloc[-1]
                                pre_close = float(last.get("close", 0))
                                prev_d = cur - timedelta(days=1)
                                while prev_d.weekday() >= 5:
                                    prev_d -= timedelta(days=1)
                                try:
                                    prev_df = df.loc[prev_d.isoformat()]
                                    pre_close = float(prev_df.iloc[-1]["close"]) if not prev_df.empty else 0
                                except Exception:
                                    pass
                                close = float(last["close"])
                                pct = round((close - pre_close) / pre_close * 100, 2) if pre_close > 0 else 0
                                rows_by_date[td_str][code] = {"close": close, "pct_chg": pct}
                        except Exception:
                            continue
                    if not rows_by_date[td_str]:
                        del rows_by_date[td_str]
                    cur += timedelta(days=1)
                if rows_by_date:
                    data_source = "local_index_1min"
            except Exception:
                pass
        # 申万行业
        try:
            sw_path = Path("f:/pythonProject/AITrade/marcus-platform/data/backtest/指数数据/sw_l1_daily.parquet")
            if sw_path.exists():
                sw_df = pd.read_parquet(sw_path)
                for td_str in list(rows_by_date.keys()):
                    try:
                        day_sw = sw_df.loc[td_str]
                        if day_sw is not None and not day_sw.empty:
                            top = day_sw.nlargest(3, "pct_change")
                            bot = day_sw.nsmallest(1, "pct_change")
                            sector_data[td_str] = {
                                "t1n": str(top.iloc[0].get("name", "")) if len(top) > 0 else "",
                                "t1p": float(top.iloc[0].get("pct_change", 0)) if len(top) > 0 else 0,
                                "t2n": str(top.iloc[1].get("name", "")) if len(top) > 1 else "",
                                "t2p": float(top.iloc[1].get("pct_change", 0)) if len(top) > 1 else 0,
                                "t3n": str(top.iloc[2].get("name", "")) if len(top) > 2 else "",
                                "t3p": float(top.iloc[2].get("pct_change", 0)) if len(top) > 2 else 0,
                                "b1n": str(bot.iloc[0].get("name", "")) if len(bot) > 0 else "",
                                "b1p": float(bot.iloc[0].get("pct_change", 0)) if len(bot) > 0 else 0,
                            }
                    except Exception:
                        continue
        except Exception:
            pass
        # 写入
        if rows_by_date:
            for ri, td_str in enumerate(sorted(rows_by_date.keys()), 2):
                row = rows_by_date[td_str]
                sec = sector_data.get(td_str, {})
                ws.cell(row=ri, column=1, value=td_str)
                for ci_off, code in enumerate(target_codes):
                    d = row.get(code, {})
                    ws.cell(row=ri, column=2 + ci_off * 2, value=d.get("close", ""))
                    ws.cell(row=ri, column=3 + ci_off * 2, value=d.get("pct_chg", ""))
                ws.cell(row=ri, column=12, value=sec.get("t1n", ""))
                ws.cell(row=ri, column=13, value=sec.get("t1p", ""))
                ws.cell(row=ri, column=14, value=sec.get("t2n", ""))
                ws.cell(row=ri, column=15, value=sec.get("t2p", ""))
                ws.cell(row=ri, column=16, value=sec.get("t3n", ""))
                ws.cell(row=ri, column=17, value=sec.get("t3p", ""))
                ws.cell(row=ri, column=18, value=sec.get("b1n", ""))
                ws.cell(row=ri, column=19, value=sec.get("b1p", ""))
        else:
            ws.cell(row=2, column=1, value=f"暂无指数数据 (data_source={data_source})")
        _auto_width(ws)
        # 标注数据源
        if data_source:
            ws.cell(row=len(rows_by_date) + 3, column=1,
                    value=f"数据源: {data_source}").font = Font(italic=True, color="666666")

        # ── Sheet 7: 策略报告 (Markdown 转纯文本) ──
        ws = wb.create_sheet("策略报告")
        ws.column_dimensions["A"].width = 110
        # 复用 strategy-report 端点的内容生成
        from sqlalchemy import func as sqlfunc
        total_trades = db.query(sqlfunc.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        buy_count = db.query(sqlfunc.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "buy").scalar() or 0
        sell_count = db.query(sqlfunc.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell").scalar() or 0
        total_comm = db.query(sqlfunc.coalesce(sqlfunc.sum(BacktestTrade.commission), 0))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        total_stamp = db.query(sqlfunc.coalesce(sqlfunc.sum(BacktestTrade.stamp_tax), 0))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        total_trans = db.query(sqlfunc.coalesce(sqlfunc.sum(BacktestTrade.transfer_fee), 0))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        sell_profit = db.query(sqlfunc.coalesce(sqlfunc.sum(BacktestTrade.profit), 0))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell").scalar() or 0
        wins = db.query(sqlfunc.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell",
                    BacktestTrade.profit > 0).scalar() or 0
        losses = db.query(sqlfunc.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell",
                    BacktestTrade.profit < 0).scalar() or 0
        denom = wins + losses
        win_rate = round(wins / denom * 100, 2) if denom > 0 else 0

        report_lines = [
            f"策略逻辑与参数报告 — {task.name}",
            "",
            f"任务ID: {task_id[:8]}",
            f"回测区间: {task.start_date} ~ {task.end_date}",
            f"初始资金: ¥{task.initial_capital:,.0f}",
            f"包含创业板: {'是' if getattr(task, 'include_chinext', False) else '否'}",
            "",
            "一、选股范围",
            "- 股票池: 全 A 股沪深两市 (排除北交所 BJ)",
            f"- 是否含创业板 (300/301 开头): {'含' if getattr(task, 'include_chinext', False) else '不含'}",
            "- 数据源: 本地 stock_basic_data.parquet (5332 只) + Tushare daily 日线",
            "- 单票仓位上限: 总资产 15% (Pi 风控规则)",
            "- 建仓时单笔金额: 不超过可用资金 40% (market_scan_1 阶段)",
            "",
            "二、买入策略",
            "- 建仓窗口: 早盘 09:35 / 09:53 / 10:35 / 13:35 (4 次扫描)",
            "- 信号源: DeepSeek LLM 解读盘中市场数据 → 产业链建仓计划",
            "- 三层过滤 (check_entry_filters):",
            "  · Layer1 技术面: MA5 > MA20 (多排) / MACD 金叉 / RSR ≥ 0.8 / 资金效率 ≥ 5%",
            "  · Layer2 主力行为: 5日主力净流入 > 0 (Tushare 超大单+大单净额累计)",
            "  · Layer3 超买过滤: RSI6 < 90 / KDJ-J < 110",
            "- 买入确认规则 (按涨幅):",
            "  · 涨幅 ≤ 2%: 直接入场",
            "  · 2% < 涨幅 ≤ 8%: 等 2-3 分钟,量比 > 1.5 才入场",
            "  · 涨幅 > 8%: 放弃 (不追涨)",
            "- T+1 规则: 买入当日不可卖出 (券商真实规则)",
            "",
            "三、卖出策略",
            "- 卖出窗口: 09:35 / 09:53 / 10:35 / 13:35 / 14:30 (尾盘强制检查)",
            "- 14:30 尾盘规则: 只卖不买,逐只检查持仓,止损触发则卖出",
            "- 动态止损 (大盘感知 + 振幅因子):",
            "  · 阈值 = max(f(大盘涨跌), 近5日日均振幅 × 0.4)",
            "  · 大盘跌 > 2% → 收紧至 -1.5% / 大盘震荡 → -2% / 大盘涨 → 放宽",
            "- 相对弱势止损: 大盘跌>2% 且个股跌幅 - 大盘跌幅 < -3pp → 强审",
            "- FIFO 成本: 同标多次买入按先进先出匹配卖出",
            "",
            "四、风控参数",
            "- 总仓位上限: 60% (green 立场) / 50% (yellow) / 20% (red 极端流出)",
            "- 单票仓位上限: 15% (核心仓 25% 仅在 3+ 指标共振+板块龙头+主力净流入)",
            "- 手续费: 买入 0.05% / 卖出 0.05% (佣金) + 0.1% (印花税) + 0.001% (沪市过户费)",
            "- 不交易清单: ST/*ST/停牌/北交所/上市 < 60 日",
            "",
            "五、回测执行统计",
            f"- 总交易笔数: {total_trades} (买入 {buy_count} / 卖出 {sell_count})",
            f"- 单笔胜率 (盈利 sell / 总 sell): {win_rate}% (胜 {wins} / 负 {losses})",
            f"- 累计实现毛盈亏: ¥{float(sell_profit):,.0f}",
            f"- 累计手续费: ¥{float(total_comm):,.0f}",
            f"- 累计印花税: ¥{float(total_stamp):,.0f}",
            f"- 累计过户费: ¥{float(total_trans):,.0f}",
            "",
        ]
        if metrics:
            report_lines += ["六、月度绩效", "月份 | 月度收益% | 交易笔数 | 胜率% | 月内最大回撤%"]
            for m in metrics:
                report_lines.append(
                    f"{m.month} | {m.return_pct:+.2f} | {m.trades_count} | "
                    f"{m.win_rate:.1f} | {m.max_drawdown:.2f}"
                )
            report_lines.append("")
        report_lines += [
            "七、反未来函数与真实性保障",
            "",
            "【数据时效】",
            "- 分钟行情: 仅取到当前 phase_time (含 +0~3 分钟 jitter) 的快照, 盘中期回退 pre_close",
            "- 资金流(个股): B2 成交额加权缩放 (下限 0.15) + 分钟K线方向修正 (±30%), 5/10日累计取历史交易日",
            "- 资金流(概念/行业/大盘): B2 成交额加权缩放, 9:30 后准入, K线蜡烛 bias 检测(强卖压→0.5x打折)",
            "- 涨跌家数: 盘中阶段只显示前日收盘数据 (防 09:35 看到当日收盘结论)",
            "- 大盘指数: 盘前用前一日收盘 / 盘中用当日 OPEN (index_1min分钟K线) / 盘后用全量日线",
            "- 盘中实时板块: get_realtime_sector_pct 基于本地指数分钟K线 (10个中证一级行业+287主题), 0~3min延迟",
            "",
            "【执行真实性】",
            "- 滑点模型: 基础0.03%(~1 tick) + 量冲击(最高0.3%) + 涨跌停加成(±3%), 总封顶5%, 买入上浮/卖出下浮",
            "- 盘中最大回撤: 用 stock_1min 找当日每只持仓最低价, 计算最坏情况权益 (非收盘价抹平日内波动)",
            "- 交易成本: 买入0.05%手续费, 卖出0.15%(含0.1%印花税+0.05%手续费), 已计入盈亏",
            "",
            "【风控保障】",
            "- T+1 强制: 当日买入当日不可卖 (PaperTradingEngine + BacktestPaperEngine._is_t1_locked 双重拦截)",
            "- T+1 持久化: 后端重启时从 PG backtest_trades 恢复 T+1 锁定状态",
            "- 总回撤 >= 5% 硬熔断: 代码层拦截所有买入, 仅允许止损卖出",
        ]
        for ri, line in enumerate(report_lines, 1):
            c = ws.cell(row=ri, column=1, value=line)
            if ri == 1:
                c.font = title_font
            elif line.startswith(("一、", "二、", "三、", "四、", "五、", "六、", "七、")):
                c.font = Font(bold=True, color="2C3E50", size=12)

        # ── 输出 ──
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = task.name or f"backtest_{task_id[:8]}"
        utf8_name = f"backtest_{safe_name}_{task_id[:8]}.xlsx"
        # 返回 (bytes, filename, data_source) 给外层 async 函数
        return buf.getvalue(), utf8_name, data_source

    finally:
        db.close()


# ─────────────────────────────────────────────────────────────
# 第二类：逐笔交易明细导出 (trades-csv)
# ─────────────────────────────────────────────────────────────

@router.get("/{task_id}/trades-csv")
async def export_trades_csv(task_id: str):
    """导出逐笔交易明细 (含分钟级时分、信号价、实际价、印花税、过户费、滑点、净盈亏)
    用途: 评估真假交易、信号滑点、手续费吞噬、胜率/盈亏比
    """
    from io import StringIO
    import csv
    from urllib.parse import quote

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        trades = (
            db.query(BacktestTrade)
            .filter(BacktestTrade.task_id == task_id)
            .order_by(BacktestTrade.trade_date, BacktestTrade.created_at)
            .all()
        )
        if not trades:
            raise HTTPException(404, "该任务暂无交易记录")

        buf = StringIO()
        writer = csv.writer(buf)
        # 表头 - 完整的逐笔交易维度
        writer.writerow([
            "trade_date", "phase_time", "datetime",       # 日期 + 时分 + 合成时分秒
            "symbol", "stock_name", "direction",            # 代码/名称/方向
            "signal_price", "actual_price", "slippage_pct",  # 信号价/实际价/滑点
            "volume", "amount",                              # 数量/金额
            "commission", "stamp_tax", "transfer_fee",       # 三项费用
            "fee_total", "fee_ratio_pct",                    # 费用合计 + 费率(占金额%)
            "profit", "profit_pct", "net_profit",            # 毛盈亏/盈亏比/净盈亏
            "t0_violation", "t0_note",                        # T+0 违规标记
            "reason",                                        # 决策理由
        ])
        for t in trades:
            td = t.trade_date.isoformat() if t.trade_date else ""
            pt = t.phase_time or ""
            dt = f"{td} {pt}:00" if pt else td
            amount = float(t.amount or 0)
            commission = float(t.commission or 0)
            stamp = float(t.stamp_tax or 0)
            transfer = float(t.transfer_fee or 0)
            fee_total = round(commission + transfer, 2)  # 卖出时 commission 已含印花税
            fee_ratio = round(fee_total / amount * 100, 4) if amount > 0 else 0
            writer.writerow([
                td, pt, dt,
                t.symbol, t.stock_name or "", t.direction,
                round(float(t.signal_price or 0), 3),
                round(float(t.actual_price or 0), 3),
                round(float(t.slippage_pct or 0), 4),
                int(t.volume or 0),
                round(amount, 2),
                round(commission, 2),
                round(stamp, 2),
                round(transfer, 2),
                fee_total, fee_ratio,
                round(float(t.profit or 0), 2),
                round(float(t.profit_pct or 0), 4),
                round(float(t.net_profit or 0), 2),
                "✓" if getattr(t, "is_t0_violation", False) else "",
                (t.t0_violation_note or "").replace("\n", " "),
                (t.reason or "").replace("\n", " "),
            ])

        safe_name = task.name or f"backtest_{task_id[:8]}"
        ascii_name = f"trades_{task_id[:8]}.csv"
        utf8_name = f"trades_{safe_name}_{task_id[:8]}.csv"
        content_disposition = (
            f"attachment; filename=\"{ascii_name}\"; "
            f"filename*=UTF-8''{quote(utf8_name)}"
        )
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": content_disposition,
                "Cache-Control": "no-cache",
            },
        )
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────
# 第三类：每日持仓明细导出 (positions-csv)
# ─────────────────────────────────────────────────────────────

@router.get("/{task_id}/positions-csv")
async def export_positions_csv(task_id: str):
    """导出每日收盘后持仓清单 (日频切片)
    用途: 单票集中度、仓位分散度、个股回撤同步性
    """
    from io import StringIO
    import csv
    from urllib.parse import quote

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        positions = (
            db.query(BacktestPosition)
            .filter(BacktestPosition.task_id == task_id)
            .order_by(BacktestPosition.trade_date, BacktestPosition.symbol)
            .all()
        )
        if not positions:
            raise HTTPException(404, "该任务暂无持仓快照")

        # 顺便取每日总资产/可用现金 (单票占比需要)
        equity = {
            e.trade_date: e
            for e in db.query(BacktestEquitySnapshot)
            .filter(BacktestEquitySnapshot.task_id == task_id)
            .all()
        }

        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "date", "symbol", "stock_name", "volume", "avg_cost", "current_price",
            "market_value", "float_pnl", "float_pnl_pct",
            "total_asset", "position_ratio_pct",  # 单票占总资产比例
        ])
        # 持仓快照里只有 symbol 没名称,回查 name 缓存(用本地 provider)
        try:
            from app.services.local_data_provider import local_data
        except Exception:
            local_data = None

        for p in positions:
            eq = equity.get(p.trade_date)
            total_asset = float(eq.total_asset) if eq and eq.total_asset else 0
            ratio = round(p.market_value / total_asset * 100, 4) \
                if total_asset > 0 and p.market_value else 0
            # 名称: 优先用 stock_name(无则查)
            name = ""
            if local_data:
                try:
                    name = local_data.get_stock_name(p.symbol) or ""
                except Exception:
                    name = ""
            writer.writerow([
                p.trade_date.isoformat() if p.trade_date else "",
                p.symbol, name,
                int(p.volume or 0),
                round(float(p.avg_cost or 0), 3),
                round(float(p.current_price or 0), 3),
                round(float(p.market_value or 0), 2),
                round(float(p.float_pnl or 0), 2),
                round(float(p.float_pnl_pct or 0), 4),
                round(total_asset, 2),
                ratio,
            ])

        safe_name = task.name or f"backtest_{task_id[:8]}"
        ascii_name = f"positions_{task_id[:8]}.csv"
        utf8_name = f"positions_{safe_name}_{task_id[:8]}.csv"
        content_disposition = (
            f"attachment; filename=\"{ascii_name}\"; "
            f"filename*=UTF-8''{quote(utf8_name)}"
        )
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": content_disposition,
                "Cache-Control": "no-cache",
            },
        )
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────
# 第一类：策略逻辑与参数报告 (strategy-report, 返回 Markdown)
# ─────────────────────────────────────────────────────────────

@router.get("/{task_id}/strategy-report")
async def get_strategy_report(task_id: str):
    """返回策略逻辑与参数摘要 (Markdown 文本)
    用途: 定性评估选股范围、买卖规则、风控纪律
    """
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        # 月度绩效
        metrics = (
            db.query(BacktestMonthlyMetric)
            .filter(BacktestMonthlyMetric.task_id == task_id)
            .order_by(BacktestMonthlyMetric.month)
            .all()
        )
        if not metrics:
            metrics = _compute_monthly_metrics_from_db(db, task_id)

        # 交易汇总
        from sqlalchemy import func
        total_trades = db.query(func.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        buy_count = db.query(func.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "buy").scalar() or 0
        sell_count = db.query(func.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell").scalar() or 0
        total_commission = db.query(func.coalesce(func.sum(BacktestTrade.commission), 0))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        total_stamp = db.query(func.coalesce(func.sum(BacktestTrade.stamp_tax), 0))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        total_transfer = db.query(func.coalesce(func.sum(BacktestTrade.transfer_fee), 0))\
            .filter(BacktestTrade.task_id == task_id).scalar() or 0
        # 胜率/盈亏比
        sell_profit_sum = db.query(func.coalesce(func.sum(BacktestTrade.profit), 0))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell").scalar() or 0
        wins = db.query(func.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell",
                    BacktestTrade.profit > 0).scalar() or 0
        losses = db.query(func.count(BacktestTrade.id))\
            .filter(BacktestTrade.task_id == task_id,
                    BacktestTrade.direction == "sell",
                    BacktestTrade.profit < 0).scalar() or 0
        denom = wins + losses
        win_rate = round(wins / denom * 100, 2) if denom > 0 else 0

        lines = []
        lines.append(f"# 策略逻辑与参数报告 — {task.name}")
        lines.append("")
        lines.append(f"- **任务ID**: `{task_id[:8]}`")
        lines.append(f"- **回测区间**: {task.start_date} ~ {task.end_date}")
        lines.append(f"- **初始资金**: ¥{task.initial_capital:,.0f}")
        lines.append(f"- **包含创业板**: {'是' if getattr(task, 'include_chinext', False) else '否'}")
        lines.append(f"- **状态**: {task.status}")
        lines.append("")

        lines.append("## 一、选股范围")
        lines.append("")
        lines.append("- **股票池**: 全 A 股沪深两市 (排除北交所 BJ)")
        lines.append(f"- **是否含创业板 (300/301 开头)**: {'含' if getattr(task, 'include_chinext', False) else '不含'}")
        lines.append("- **数据源**: 本地 stock_basic_data.parquet (5332 只) + Tushare daily 日线")
        lines.append("- **单票仓位上限**: 总资产 15% (Pi 风控规则)")
        lines.append("- **建仓时单笔金额**: 不超过可用资金 40% (market_scan_1 阶段)")
        lines.append("")

        lines.append("## 二、买入策略")
        lines.append("")
        lines.append("- **建仓窗口**: 早盘 09:35 / 09:53 / 10:35 / 13:35 (4 次扫描)")
        lines.append("- **信号源**: DeepSeek LLM 解读盘中市场数据 → 产业链建仓计划")
        lines.append("- **三层过滤** (check_entry_filters):")
        lines.append("  - **Layer1 技术面**: MA5 > MA20 (多排) / MACD 金叉 / RSR ≥ 0.8 / 资金效率 ≥ 5%")
        lines.append("  - **Layer2 主力行为**: 5日主力净流入 > 0 (Tushare 超大单+大单净额累计)")
        lines.append("  - **Layer3 超买过滤**: RSI6 < 90 / KDJ-J < 110")
        lines.append("- **买入确认规则** (按涨幅):")
        lines.append("  - 涨幅 ≤ 2%: 直接入场")
        lines.append("  - 2% < 涨幅 ≤ 8%: 等 2-3 分钟，量比 > 1.5 才入场")
        lines.append("  - 涨幅 > 8%: 放弃（不追涨）")
        lines.append("- **T+1 规则**: 买入当日不可卖出（券商真实规则）")
        lines.append("")

        lines.append("## 三、卖出策略")
        lines.append("")
        lines.append("- **卖出窗口**: 09:35 / 09:53 / 10:35 / 13:35 / 14:30 (尾盘强制检查)")
        lines.append("- **14:30 尾盘规则**: 只卖不买，逐只检查持仓，止损触发则卖出")
        lines.append("- **动态止损** (大盘感知 + 振幅因子):")
        lines.append("  - 阈值 = max(f(大盘涨跌), 近5日日均振幅 × 0.4)")
        lines.append("  - 大盘跌 > 2% → 收紧至 -1.5% / 大盘震荡 → -2% / 大盘涨 → 放宽")
        lines.append("- **相对弱势止损**: 大盘跌>2% 且个股跌幅 - 大盘跌幅 < -3pp → 强审")
        lines.append("- **FIFO 成本**: 同标多次买入按先进先出匹配卖出")
        lines.append("")

        lines.append("## 四、风控参数")
        lines.append("")
        lines.append("- **总仓位上限**: 60% (green 立场) / 50% (yellow) / 20% (red 极端流出)")
        lines.append("- **单票仓位上限**: 15% (核心仓 25% 仅在 3+ 指标共振+板块龙头+主力净流入)")
        lines.append("- **手续费**: 买入 0.05% / 卖出 0.05% (佣金) + 0.1% (印花税) + 0.001% (沪市过户费)")
        lines.append("- **不交易清单**: ST/*ST/停牌/北交所/上市 < 60 日")
        lines.append("")

        lines.append("## 五、回测执行统计")
        lines.append("")
        lines.append(f"- **总交易笔数**: {total_trades} (买入 {buy_count} / 卖出 {sell_count})")
        lines.append(f"- **单笔胜率** (盈利 sell / 总 sell): **{win_rate}%** (胜 {wins} / 负 {losses})")
        sell_pct = round(sell_profit_sum, 2) if sell_profit_sum else 0
        lines.append(f"- **累计实现毛盈亏**: ¥{sell_pct:,.0f}")
        lines.append(f"- **累计手续费**: ¥{round(float(total_commission), 2):,.0f}")
        lines.append(f"- **累计印花税**: ¥{round(float(total_stamp), 2):,.0f}")
        lines.append(f"- **累计过户费**: ¥{round(float(total_transfer), 2):,.0f}")
        lines.append("")

        if metrics:
            lines.append("## 六、月度绩效")
            lines.append("")
            lines.append("| 月份 | 月度收益 | 交易笔数 | 胜率 | 月内最大回撤 |")
            lines.append("|------|---------:|--------:|-----:|------------:|")
            for m in metrics:
                lines.append(f"| {m.month} | {m.return_pct:+.2f}% | "
                             f"{m.trades_count} | {m.win_rate:.1f}% | "
                             f"{m.max_drawdown:.2f}% |")
            lines.append("")

        lines.append("## 七、反未来函数与真实性保障")
        lines.append("")
        lines.append("### 数据时效")
        lines.append("")
        lines.append("- **分钟行情**: 仅取到当前 phase_time (含 +0~3 分钟 jitter) 的快照, 盘中期回退 pre_close")
        lines.append("- **资金流(个股)**: B2 成交额加权缩放 (下限 0.15) + 分钟K线方向修正 (±30%), 5/10日累计取历史交易日")
        lines.append("- **资金流(概念/行业/大盘)**: B2 成交额加权缩放, 9:30 后准入, K线蜡烛 bias 检测(强卖压→0.5x打折)")
        lines.append("- **涨跌家数**: 盘中阶段只显示前日收盘数据 (防 09:35 看到当日收盘结论)")
        lines.append("- **大盘指数**: 盘前用前一日收盘 / 盘中用当日 OPEN (index_1min分钟K线) / 盘后用全量日线")
        lines.append("- **盘中实时板块**: `get_realtime_sector_pct` 基于本地指数分钟K线 (10个中证一级行业+287主题), 0~3min延迟")
        lines.append("")
        lines.append("### 执行真实性")
        lines.append("")
        lines.append("- **滑点模型**: 基础0.03%(~1 tick) + 量冲击(最高0.3%) + 涨跌停加成(±3%), 总封顶5%, 买入上浮/卖出下浮")
        lines.append("- **盘中最大回撤**: 用 stock_1min 找当日每只持仓最低价, 计算最坏情况权益 (非收盘价抹平日内波动)")
        lines.append("- **交易成本**: 买入0.05%手续费, 卖出0.15%(含0.1%印花税+0.05%手续费), 已计入盈亏")
        lines.append("")
        lines.append("### 风控保障")
        lines.append("")
        lines.append("- **T+1 强制**: 当日买入当日不可卖 (PaperTradingEngine + BacktestPaperEngine 双重拦截)")
        lines.append("- **T+1 持久化**: 后端重启时从 PG backtest_trades 恢复 T+1 锁定状态")
        lines.append("- **总回撤熔断**: 总回撤 >= 5% 代码层硬拦截所有买入, 仅允许止损卖出")

        return {
            "task_id": task_id,
            "markdown": "\n".join(lines),
            "stats": {
                "total_trades": total_trades,
                "buy_count": buy_count,
                "sell_count": sell_count,
                "win_rate": win_rate,
                "wins": wins, "losses": losses,
                "total_commission": round(float(total_commission), 2),
                "total_stamp_tax": round(float(total_stamp), 2),
                "total_transfer_fee": round(float(total_transfer), 2),
            },
        }
    finally:
        db.close()


# ─────────────────────────────────────────────────────────────
# 第四类：基准指数 + 市场背景数据 (index-csv, 来自 sandbox/indices 已存日志)
# ─────────────────────────────────────────────────────────────

@router.get("/{task_id}/index-csv")
async def export_index_csv(task_id: str):
    """导出回测期间主要指数日线 (沪深300/中证500/创业板/科创50/上证)

    数据源: Tushare index_daily (通过 sandbox/indices 缓存) + 本地 index_1min (兜底)
    """
    from io import StringIO
    import csv
    from urllib.parse import quote
    from pathlib import Path

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")

        # 5 大指数
        target_codes = ["000001.SH", "399001.SZ", "000300.SH", "399006.SZ", "000688.SH"]
        name_map = {
            "000001.SH": "上证指数", "399001.SZ": "深证成指",
            "000300.SH": "沪深300", "399006.SZ": "创业板指", "000688.SH": "科创50",
        }

        rows_by_date: Dict[str, Dict[str, dict]] = {}
        data_source = ""

        # 方案 A: 从 BacktestDailyLog 的 metadata_json 提取 (若 Pi 调过 sandbox/indices)
        try:
            logs = (
                db.query(BacktestDailyLog)
                .filter(BacktestDailyLog.task_id == task_id)
                .filter(BacktestDailyLog.event_type.in_(["scan_report", "pi_trade"]))
                .filter(BacktestDailyLog.metadata_json.isnot(None))
                .all()
            )
            for log in logs:
                meta = log.metadata_json or {}
                # 1) account_summary 里可能嵌套 indices (旧版)
                if not meta.get("indices"):
                    continue
                td = log.trade_date.isoformat() if log.trade_date else ""
                if not td or td not in rows_by_date:
                    rows_by_date[td] = {}
                for idx in meta.get("indices", []):
                    code = idx.get("ts_code") or idx.get("code")
                    if code in target_codes:
                        rows_by_date[td][code] = {
                            "close": float(idx.get("close") or idx.get("current_price") or 0),
                            "pct_chg": float(idx.get("change_pct") or idx.get("pct_chg") or 0),
                            "open": float(idx.get("open") or 0),
                            "high": float(idx.get("high") or 0),
                            "low": float(idx.get("low") or 0),
                        }
            if rows_by_date:
                data_source = "backtest_daily_log_cache"
        except Exception:
            pass

        # 方案 B: 直接读本地 index_1min 推算 (fallback, 数据可能不完整)
        if not rows_by_date:
            try:
                base = Path("f:/pythonProject/AITrade/marcus-platform/data/backtest/指数数据/index_1min")
                from datetime import date as dt_date, timedelta
                start = task.start_date or dt_date(2026, 1, 1)
                end = task.end_date or dt_date.today()
                cur = start
                while cur <= end:
                    td_str = cur.isoformat()
                    rows_by_date[td_str] = {}
                    for code in target_codes:
                        fp = base / f"{code}.parquet"
                        if not fp.exists():
                            continue
                        try:
                            df = pd.read_parquet(fp)
                            # 取该日最后一条 (收盘)
                            day_df = df.loc[td_str] if td_str in df.index.get_level_values(0) else None
                            if day_df is not None and not day_df.empty:
                                last = day_df.iloc[-1]
                                first = day_df.iloc[0]
                                pre_close = float(last.get("close", 0))  # 兜底
                                # 取昨日最后一条
                                prev_d = cur - timedelta(days=1)
                                while prev_d.weekday() >= 5:
                                    prev_d -= timedelta(days=1)
                                try:
                                    prev_df = df.loc[prev_d.isoformat()]
                                    pre_close = float(prev_df.iloc[-1]["close"]) if not prev_df.empty else 0
                                except Exception:
                                    pass
                                close = float(last["close"])
                                pct = round((close - pre_close) / pre_close * 100, 2) if pre_close > 0 else 0
                                rows_by_date[td_str][code] = {
                                    "close": close,
                                    "pct_chg": pct,
                                    "open": float(first["open"]),
                                    "high": float(day_df["high"].max()),
                                    "low": float(day_df["low"].min()),
                                }
                        except Exception:
                            continue
                    # 清理空日
                    if not rows_by_date[td_str]:
                        del rows_by_date[td_str]
                    cur += timedelta(days=1)
                if rows_by_date:
                    data_source = "local_index_1min"
            except Exception as e:
                pass

        # ── 行业/概念板块背景 (来自 sw_l1_daily 当日均涨跌幅) ──
        sector_data: Dict[str, Dict[str, float]] = {}
        try:
            sw_path = Path("f:/pythonProject/AITrade/marcus-platform/data/backtest/指数数据/sw_l1_daily.parquet")
            if sw_path.exists():
                sw_df = pd.read_parquet(sw_path)
                for td_str in list(rows_by_date.keys()):
                    try:
                        day_sw = sw_df.loc[td_str]
                        if day_sw is not None and not day_sw.empty:
                            top = day_sw.nlargest(3, "pct_change")
                            bot = day_sw.nsmallest(3, "pct_change")
                            sector_data[td_str] = {
                                "top1_name": str(top.iloc[0].get("name", "")) if len(top) > 0 else "",
                                "top1_pct": float(top.iloc[0].get("pct_change", 0)) if len(top) > 0 else 0,
                                "top2_name": str(top.iloc[1].get("name", "")) if len(top) > 1 else "",
                                "top2_pct": float(top.iloc[1].get("pct_change", 0)) if len(top) > 1 else 0,
                                "top3_name": str(top.iloc[2].get("name", "")) if len(top) > 2 else "",
                                "top3_pct": float(top.iloc[2].get("pct_change", 0)) if len(top) > 2 else 0,
                                "bot1_name": str(bot.iloc[0].get("name", "")) if len(bot) > 0 else "",
                                "bot1_pct": float(bot.iloc[0].get("pct_change", 0)) if len(bot) > 0 else 0,
                            }
                    except Exception:
                        continue
        except Exception:
            pass

        if not rows_by_date:
            raise HTTPException(404, "该任务无指数数据 (回测未调用 sandbox/indices,本地 index_1min 也不覆盖)")

        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            "date",
            "sh_close", "sh_pct",
            "sz_close", "sz_pct",
            "hs300_close", "hs300_pct",
            "gem_close", "gem_pct",
            "star50_close", "star50_pct",
            "sector_top1", "sector_top1_pct",
            "sector_top2", "sector_top2_pct",
            "sector_top3", "sector_top3_pct",
            "sector_bot1", "sector_bot1_pct",
        ])
        for td_str in sorted(rows_by_date.keys()):
            row = rows_by_date[td_str]
            sec = sector_data.get(td_str, {})
            def _g(code):
                d = row.get(code, {})
                return d.get("close", ""), d.get("pct_chg", "")
            sh_c, sh_p = _g("000001.SH")
            sz_c, sz_p = _g("399001.SZ")
            hs_c, hs_p = _g("000300.SH")
            gem_c, gem_p = _g("399006.SZ")
            star_c, star_p = _g("000688.SH")
            writer.writerow([
                td_str,
                sh_c, sh_p, sz_c, sz_p, hs_c, hs_p, gem_c, gem_p, star_c, star_p,
                sec.get("top1_name", ""), sec.get("top1_pct", ""),
                sec.get("top2_name", ""), sec.get("top2_pct", ""),
                sec.get("top3_name", ""), sec.get("top3_pct", ""),
                sec.get("bot1_name", ""), sec.get("bot1_pct", ""),
            ])

        safe_name = task.name or f"backtest_{task_id[:8]}"
        ascii_name = f"index_{task_id[:8]}.csv"
        utf8_name = f"index_{safe_name}_{task_id[:8]}.csv"
        content_disposition = (
            f"attachment; filename=\"{ascii_name}\"; "
            f"filename*=UTF-8''{quote(utf8_name)}"
        )
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": content_disposition,
                "Cache-Control": "no-cache",
                "X-Data-Source": data_source,
            },
        )
    finally:
        db.close()


@router.get("/{task_id}/prompt-snapshot/{log_id}")
async def get_prompt_snapshot(task_id: str, log_id: int):
    """获取指定日志条目的完整 prompt 快照和回复"""
    db = SessionLocal()
    try:
        log = (
            db.query(BacktestDailyLog)
            .filter(BacktestDailyLog.id == log_id, BacktestDailyLog.task_id == task_id)
            .first()
        )
        if not log:
            raise HTTPException(404, "日志不存在")

        meta = log.metadata_json or {}
        return {
            "id": log.id,
            "trade_date": log.trade_date.isoformat(),
            "phase": log.phase,
            "phase_time": log.phase_time,
            "event_type": log.event_type,
            "prompt_snapshot": meta.get("prompt_snapshot"),
            "full_reply": meta.get("full_reply"),
        }
    finally:
        db.close()


@router.delete("/{task_id}")
async def delete_backtest(task_id: str):
    """删除回测任务及其所有关联数据"""
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")
        if task.status == "running":
            raise HTTPException(400, "运行中的任务无法删除，请先取消")

        db.delete(task)
        db.commit()
        return {"success": True, "message": "回测任务已删除"}
    finally:
        db.close()


# ── 后台执行 ──

async def _run_backtest_async(task_id: str, on_event_callback=None):
    """在后台异步运行回测"""
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            return
        start_date = task.start_date
        end_date = task.end_date
        initial_capital = task.initial_capital
        include_chinext = getattr(task, "include_chinext", True) or True
        model_name = getattr(task, "model_name", "deepseek-v4-pro") or "deepseek-v4-pro"
        thinking_level = getattr(task, "thinking_level", "high") or "high"
    finally:
        db.close()

    await backtest_engine.run(
        task_id=task_id,
        start_date=start_date,
        end_date=end_date,
        initial_capital=initial_capital,
        include_chinext=include_chinext,
        model_name=model_name,
        thinking_level=thinking_level,
        on_event=on_event_callback,
    )


# ══════════════════════════════════════════════════════════
# 沙盒账户 API（供 Pi Server backtest 模式调用）
# ══════════════════════════════════════════════════════════

class SandboxOrderRequest(BaseModel):
    symbol: str = Field(..., description="股票代码")
    direction: str = Field(..., description="buy/sell")
    price: float = Field(..., description="委托价格（仅参考，实际成交价由 phase_time 决定）")
    volume: int = Field(..., description="股数")
    reason: str = Field("", description="交易理由")
    phase_time: Optional[str] = Field(None, description="Pi 调用时刻 HH:MM（用于反查分钟成交价）")


@router.get("/{task_id}/sandbox/account")
async def get_sandbox_account(task_id: str, trade_date: str = Query(None)):
    """获取沙盒账户状态（优先从 PaperTradingEngine 实时查，回退 PG）"""
    from app.services.backtest_engine import backtest_engine
    engine = backtest_engine._engines.get(task_id)
    if engine:
        acc = engine.get_account()
        acc["trade_date"] = trade_date
        return acc
    # 回退 PG
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")
        eq = (db.query(BacktestEquitySnapshot).filter(
            BacktestEquitySnapshot.task_id == task_id
        ).order_by(BacktestEquitySnapshot.trade_date.desc()).first())
        cap = task.initial_capital
        if eq:
            return {"total_asset": eq.total_asset, "available_cash": eq.available_cash,
                    "position_value": eq.position_value, "initial_capital": cap,
                    "return_pct": round((eq.total_asset/cap-1)*100, 2)}
        return {"total_asset": cap, "available_cash": cap, "position_value": 0,
                "initial_capital": cap, "return_pct": 0}
    finally:
        db.close()


@router.get("/{task_id}/sandbox/positions")
async def get_sandbox_positions(task_id: str, trade_date: str = Query(None)):
    """获取沙盒持仓（供 Pi 回测工具调用）
    包含 T+1 锁定状态：locked / last_buy_date / unlock_date / reason
    """
    from app.services.backtest_engine import backtest_engine
    db = SessionLocal()
    try:
        q = db.query(BacktestPosition).filter(BacktestPosition.task_id == task_id)
        if trade_date:
            q = q.filter(BacktestPosition.trade_date == date.fromisoformat(trade_date))
        else:
            # 获取最近有持仓的日期
            latest = (
                db.query(BacktestPosition)
                .filter(BacktestPosition.task_id == task_id)
                .order_by(BacktestPosition.trade_date.desc())
                .first()
            )
            if latest:
                q = q.filter(BacktestPosition.trade_date == latest.trade_date)

        positions = q.all()

        # 从实时引擎取 T+1 状态（_last_buy_date 在内存里，最准确）
        engine = backtest_engine._engines.get(task_id)
        result_positions = []
        for p in positions:
            pos_dict = {
                "symbol": p.symbol,
                "volume": p.volume,
                "avg_cost": p.avg_cost,
                "current_price": p.current_price,
                "market_value": p.market_value,
                "float_pnl": p.float_pnl,
                "float_pnl_pct": p.float_pnl_pct,
            }
            if engine:
                # 归一化 symbol 为 BacktestPaperEngine 用的格式
                s = p.symbol
                engine_sym = s if s.startswith(("SH", "SZ")) else (
                    "SH" + s.split(".")[0] if s.endswith(".SH")
                    else "SZ" + s.split(".")[0] if s.endswith(".SZ")
                    else s
                )
                t1 = engine.get_t1_status(engine_sym)
                pos_dict["t1_status"] = t1
            else:
                pos_dict["t1_status"] = {"locked": False, "last_buy_date": None,
                                          "unlock_date": None, "reason": "引擎已结束"}
            result_positions.append(pos_dict)

        return {
            "positions": result_positions,
            "count": len(result_positions),
            "trade_date": trade_date,
        }
    finally:
        db.close()


@router.post("/{task_id}/sandbox/order")
async def place_sandbox_order(task_id: str, req: SandboxOrderRequest):
    """在沙盒账户中下单（通过真实 PaperTradingEngine）

    ⚠️ 反未来函数: 实际成交价不由 Pi 传的 price 决定, 而由 phase_time 对应的
    分钟行情决定 (9:35 → 9:35 价; 14:30 → 14:30 价)。这样 Pi 不能"预知"收盘价
    之后再以收盘价下单锁定胜率。
    """
    from app.services.backtest_engine import backtest_engine
    from app.core.trading.backtest_paper import BacktestPaperEngine
    from app.services.local_data_provider import local_data

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, "任务不存在")
        if task.status != "running":
            raise HTTPException(400, "任务未在运行中")

        trade_date = task.current_day or date.today()

        # 使用真实 PaperTradingEngine 执行
        engine = backtest_engine._engines.get(task_id)
        if not engine:
            engine = BacktestPaperEngine(task_id, task.initial_capital)
            backtest_engine._engines[task_id] = engine

        # ⚠️ 关键: 设置当前模拟日期, T+1 校验才能工作
        # 如果不设 _current_date, _is_t1_locked() 总是返回 False (A 股 T+1 形同虚设)
        engine.set_current_date(trade_date)

        # ── 反未来函数: 强制用 phase_time 对应的"已知"价格成交 ──
        # 策略:
        #   1. 如果传了 phase_time 且 < 16:00 (盘中期) → 走分钟数据
        #   2. 如果 phase_time >= 16:00 或 缺失 → 用当日 open (开盘价)
        #   3. 都没有 → 拒绝交易 (避免用 pre_close 暗藏未来函数)
        actual_price = req.price  # 默认值
        price_source = "client_price"
        price_warning = None

        if req.phase_time:
            try:
                hh, mm = map(int, req.phase_time.split(":")[:2])
                cur_minutes = hh * 60 + mm
            except Exception:
                cur_minutes = None
        else:
            cur_minutes = None

        if cur_minutes is not None and cur_minutes < 16 * 60:
            # 盘中期: 必须用分钟行情
            mq = local_data.get_minute_quote(req.symbol, trade_date, hh, mm)
            if mq and float(mq.get("close", 0)) > 0:
                actual_price = float(mq["close"])
                price_source = f"minute_{req.phase_time}"
            else:
                # 分钟数据缺失 → 拒绝交易, 让 Pi 用更早的 phase 重试
                return {
                    "success": False,
                    "message": f"⚠️ 拒绝成交: {req.symbol} 在 {trade_date} {req.phase_time} 无分钟数据, 无法反查真实价",
                }
        else:
            # 盘后 (16:00+): 可以用 open 价
            dq = local_data.get_daily_quote(req.symbol, trade_date)
            if dq and float(dq.get("open", 0)) > 0:
                actual_price = float(dq["open"])
                price_source = "daily_open"
            else:
                return {
                    "success": False,
                    "message": f"⚠️ 拒绝成交: {req.symbol} 在 {trade_date} 无日线数据",
                }

        # ── 滑点模型 (模拟实盘成交磨损) ──
        # 如果 Pi 传了 phase_time, 从分钟数据估算当日成交量用于量冲击
        slip_pct = 0.0
        slip_detail = ""
        try:
            if req.phase_time and cur_minutes is not None and cur_minutes < 16 * 60:
                mq = local_data.get_minute_quote(req.symbol, trade_date, hh, mm)
                if mq:
                    # 基础滑点: 0.03% (约 1 tick)
                    slip_pct = 0.0003
                    # 量冲击: 委托量 / 当日累计成交量 (上限 0.3%)
                    day_vol = float(mq.get("volume_day", 0) or 0)
                    if day_vol > 0:
                        vol_ratio = min(req.volume / day_vol, 1.0)
                        slip_pct += vol_ratio * 0.003
                    # 极端行情: 涨跌停附近加成 0.5%
                    prev_close = float(mq.get("pre_close", 0) or 0)
                    if prev_close > 0:
                        move_pct = abs(actual_price / prev_close - 1) * 100
                        if move_pct > 9.5:
                            # 涨跌停板: 大幅滑点 + 流动性风险注入
                            slip_pct += 0.03
                            slip_detail = f" 涨跌停加成+3%"
                        elif move_pct > 5:
                            slip_pct += 0.002
                            slip_detail = f" 极端波动+0.2%"
                    slip_detail = f"基础0.03%+量{vol_ratio:.1%}*0.3%{slip_detail}"
        except Exception:
            slip_pct = 0.0003  # fallback: 基础滑点

        # 买入价上浮 / 卖出价下浮
        if req.direction.lower() == "buy":
            slip_price = round(actual_price * (1 + min(slip_pct, 0.05)), 2)
        else:
            slip_price = round(actual_price * (1 - min(slip_pct, 0.05)), 2)

        result = engine.place_order(req.symbol, req.direction, slip_price, req.volume)

        # ⚠️ P0 修复: 失败订单不再写 PG (T+1 拦截 / 资金不足 / 持仓不足)
        # 旧代码: 不管 success 都写 BacktestTrade, 导致报告里出现"未成交"的幽灵交易
        if not result.get("success"):
            return {
                "success": False,
                "message": result.get("message", "下单失败"),
                "fill_price": None,
                "price_source": price_source,
            }

        # ── 计算实现盈亏（仅 sell 才有意义） ──
        # 关键: 直接从 engine 内部 trades.db 读 FIFO 算出的 profit
        # (PaperTradingEngine.match_order 内部已经按 FIFO 计算并写入 trades.profit 字段)
        # 之前用 positions[sym].avg_price 算有 3 个问题:
        #   1. T+0 短线 sell 后 avg_price 已被减仓算法改写
        #   2. 卖光后 pos 被 delete,拿不到 avg_price
        #   3. 部分卖出后 avg_price 反映的是"剩余持仓",不是被卖出的那部分
        profit = 0.0
        profit_pct = 0.0
        if req.direction.lower() == "sell" and result.get("success"):
            try:
                # 归一化为 engine 内部 symbol (SH600519 / SZ000001)
                s = req.symbol
                engine_sym = s if s.startswith(("SH", "SZ")) else (
                    "SH" + s.split(".")[0] if s.endswith(".SH")
                    else "SZ" + s.split(".")[0] if s.endswith(".SZ")
                    else s
                )
                # 拉最近一笔该 symbol 的 sell 成交 (就是我们刚下的)
                last_trades = engine._engine.get_trades(symbol=engine_sym, limit=1)
                if last_trades and last_trades[0].get("direction") in ("sell", "short", "SELL", "卖出"):
                    fifo_profit = float(last_trades[0].get("profit", 0) or 0)
                    # 真实盈亏 = FIFO profit - 印花税(0.1%) - 手续费(0.05%)
                    # 之所以减: engine 的 profit 是"金额差额",还没扣交易成本
                    sell_amount = slip_price * req.volume  # 滑点后实际成交金额
                    stamp_tax = sell_amount * 0.001
                    commission = sell_amount * 0.0005
                    profit = round(fifo_profit - stamp_tax - commission, 2)
                    if slip_price > 0:
                        # profit_pct 相对 FIFO 成本(不含交易成本,因为引擎也不扣)
                        # 简化: 用 fifo_profit / sell_amount
                        profit_pct = round(fifo_profit / sell_amount * 100, 4) if sell_amount > 0 else 0
            except Exception as e:
                logger.warning(f"sell profit 计算失败: {e}")
                profit = 0.0
                profit_pct = 0.0

        # ── 查股票名称（覆盖全 A 股） ──
        stock_name = ""
        try:
            stock_name = local_data.get_stock_name(req.symbol) or ""
        except Exception:
            stock_name = ""

        # ── 计算滑点 / 过户费 / 净盈亏 (导出逐笔明细用) ──
        trade_amount = round(slip_price * req.volume, 2)
        # 滑点%
        slip = 0.0
        if actual_price > 0:
            if req.direction.lower() == "buy":
                slip = round((slip_price / actual_price - 1) * 100, 4)
            else:
                slip = round((slip_price / actual_price - 1) * 100, 4)
        # 过户费 (沪市 0.001%, 深市不收)
        is_sh = bool(re.match(r'^(SH)?6\d{5}', req.symbol.upper()))
        transfer_fee = round(trade_amount * 0.00001, 2) if is_sh else 0.0
        # 印花税 (仅卖出, 已从 profit 中扣除, 此处记录明细)
        trade_stamp = round(trade_amount * 0.001, 2) if req.direction.lower() == "sell" else 0.0
        trade_commission = round(trade_amount * 0.0005, 2)
        # 净盈亏 = profit - 过户费 (印花税和佣金已含在 profit 中)
        net_profit = round(float(profit) - transfer_fee, 2)

        # 同步到 PG (记录滑点后实际成交价 + Pi 信号价)
        trade_record = BacktestTrade(
            task_id=task_id, trade_date=trade_date,
            symbol=req.symbol, stock_name=stock_name,
            direction=req.direction,
            price=slip_price, volume=req.volume,
            amount=trade_amount,
            commission=trade_commission,
            signal_price=round(actual_price, 2),
            actual_price=round(slip_price, 2),
            stamp_tax=trade_stamp,
            transfer_fee=transfer_fee,
            slippage_pct=round(slip, 4),
            net_profit=net_profit,
            phase_time=req.phase_time or "",
            profit=profit, profit_pct=profit_pct,
            reason=req.reason,
        )
        db.add(trade_record)
        db.commit()

        return {
            "success": result["success"],
            "message": result["message"],
            "fill_price": actual_price,
            "price_source": price_source,
        }
    finally:
        db.close()


@router.get("/{task_id}/sandbox/orders")
async def get_sandbox_orders(task_id: str,
                               symbol: str = Query(None),
                               status: str = Query(None),
                               limit: int = Query(50, ge=1, le=200)):
    """获取沙盒账户的活跃订单（回测模式专用，避免触发真实 MarcusVNPyExecutor）
    数据源: 沙盒 PaperTradingEngine 的 trades.db"""
    from app.services.backtest_engine import backtest_engine
    from app.core.trading.backtest_paper import BacktestPaperEngine

    engine = backtest_engine._engines.get(task_id)
    if not engine:
        # 任务未运行, 尝试构造 (只读, 不持久化)
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            if not task:
                return {"orders": [], "count": 0, "error": "任务不存在"}
            engine = BacktestPaperEngine(task_id, task.initial_capital)
        finally:
            db.close()

    try:
        # 从沙盒 PaperTradingEngine 读 orders 表
        all_orders = engine._engine.get_orders(symbol=symbol, status=status, limit=limit)
        return {"orders": all_orders, "count": len(all_orders)}
    except Exception as e:
        return {"orders": [], "count": 0, "error": str(e)}


@router.get("/{task_id}/sandbox/scan-report")
async def get_sandbox_scan_report(task_id: str, trade_date: str = Query(None)):
    """获取沙盒扫描报告（供 Pi 回测工具调用）"""
    db = SessionLocal()
    try:
        q = (
            db.query(BacktestDailyLog)
            .filter(BacktestDailyLog.task_id == task_id)
            .filter(BacktestDailyLog.event_type.in_(["scan_report", "pi_analysis", "pi_trade"]))
        )
        if trade_date:
            q = q.filter(BacktestDailyLog.trade_date == date.fromisoformat(trade_date))
        q = q.order_by(BacktestDailyLog.trade_date.desc(), BacktestDailyLog.id.desc()).limit(3)

        logs = q.all()
        return {
            "reports": [
                {
                    "trade_date": l.trade_date.isoformat(),
                    "phase": l.phase,
                    "event_type": l.event_type,
                    "content": l.content[:2000] if l.content else "",
                    "metadata": l.metadata_json,
                }
                for l in logs
            ],
            "count": len(logs),
        }
    finally:
        db.close()


# ══════════════════════════════════════════════════════════
# 本地数据服务 API（供 Pi 回测工具查询历史数据）
# ══════════════════════════════════════════════════════════

# A股收盘时间常量（用于"未来函数"判断）
MARKET_CLOSE_HOUR = 15  # 15:00 收盘
MARKET_CLOSE_MINUTE = 0
# 收盘后 + 1小时缓冲 (16:00 后认为当天 K 线已落盘)
MARKET_FINALIZED_HOUR = 16
MARKET_FINALIZED_MINUTE = 0


def _is_day_finalized(phase_time: str = None) -> bool:
    """判断当天 K 线是否已"落盘"（避免未来函数）
    - phase_time: Pi 调用时的时分（HH:MM），来自 [BKT:task_id|date|HH:MM] 前缀
    - 收盘 15:00 之后 + 1 小时缓冲 (16:00) → 当天 K 线视为已确认
    - 16:00 之前 → 当天视为未知，排除
    适用范围: K线 / 技术指标 / 行情快照（价格类数据）
    """
    if not phase_time:
        return False
    try:
        hh, mm = map(int, phase_time.split(":")[:2])
        cur = hh * 60 + mm
        return cur >= (MARKET_FINALIZED_HOUR * 60 + MARKET_FINALIZED_MINUTE)
    except Exception:
        return False


def _is_moneyflow_available(phase_time: str = None) -> bool:
    """判断盘中资金流数据是否可用（B2缩放场景专用）
    - 资金流数据用 B2 成交额加权缩放作为反未来函数手段
    - 不需要等收盘，开盘(9:30)后即可提供当日渐进数据
    - 不同 phase_time 返回不同的缩放权重，数据递增趋近真实值
    - 没有 phase_time → 视为盘前，返回前日数据
    适用范围: 概念/行业/大盘资金流（非价格类数据）
    """
    if not phase_time:
        return False
    try:
        hh, mm = map(int, phase_time.split(":")[:2])
        cur = hh * 60 + mm
        return cur >= (9 * 60 + 30)  # 9:30 开盘
    except Exception:
        return False


@router.get("/{task_id}/sandbox/kline/{symbol}")
async def get_sandbox_kline(task_id: str, symbol: str,
                             limit: int = Query(30, ge=1, le=120),
                             trade_date: str = Query(None),
                             phase_time: str = Query(None)):
    """返回历史日K线（从本地 parquet），Pi 的 get_daily_kline 回测路由到此
    phase_time: Pi 调用时刻 HH:MM（来自 [BKT:...] 前缀）
    - 16:00 之前: 排除当天 K 线（未来函数防护）
    - 16:00 之后: 当天 K 线视为已收盘确认，可返回
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    # 如果没有传日期，从任务中获取当前日期
    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"kline": [], "error": "无法确定回测日期"}

    end_dt = dt_date.fromisoformat(trade_date)
    start_dt = end_dt - timedelta(days=limit * 2)  # 留足余量跳过非交易日

    # ⚠️ 16:00 之前排除当天 (避免未来函数), 16:00 之后认为已收盘
    include_today = _is_day_finalized(phase_time)
    bound = end_dt if include_today else (end_dt - timedelta(days=1))

    klines = []
    current = start_dt
    while current <= bound and len(klines) < limit:
        q = local_data.get_daily_quote(symbol, current)
        if q:
            klines.append({
                "date": q["date"], "open": q["open"], "high": q["high"],
                "low": q["low"], "close": q["close"], "volume": q["volume"],
                "amount": q["amount"],
            })
        current += timedelta(days=1)

    return {
        "kline": klines[-limit:], "symbol": symbol, "count": len(klines[-limit:]),
        "include_today": include_today, "phase_time": phase_time,
    }


@router.get("/{task_id}/sandbox/technical/{symbol}")
async def get_sandbox_technical(task_id: str, symbol: str,
                                 trade_date: str = Query(None),
                                 phase_time: str = Query(None)):
    """返回技术指标（基于本地日线计算），Pi 的 get_technical 回测路由到此
    16:00 前排除当天 K 线（未来函数防护）"""
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"error": "无法确定回测日期"}

    end_dt = dt_date.fromisoformat(trade_date)
    start_dt = end_dt - timedelta(days=60)

    include_today = _is_day_finalized(phase_time)
    bound = end_dt if include_today else (end_dt - timedelta(days=1))

    closes = []
    highs = []
    lows = []
    current = start_dt
    while current <= bound:
        q = local_data.get_daily_quote(symbol, current)
        if q:
            closes.append(q["close"])
            highs.append(q["high"])
            lows.append(q["low"])
        current += timedelta(days=1)

    if len(closes) < 5:
        return {"symbol": symbol, "error": "数据不足"}

    # 简易 MACD/KDJ/RSI/MA 计算
    closes_s = pd.Series(closes)
    ema12 = closes_s.ewm(span=12).mean()
    ema26 = closes_s.ewm(span=26).mean()
    dif = ema12 - ema26
    dea = dif.ewm(span=9).mean()
    macd_bar = 2 * (dif - dea)

    # RSI(14)
    delta = closes_s.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    avg_gain = gain.ewm(alpha=1/14).mean()
    avg_loss = loss.ewm(alpha=1/14).mean()
    rs = avg_gain / avg_loss.replace(0, 1e-9)
    rsi = 100 - (100 / (1 + rs))

    # MA
    ma5 = closes_s.rolling(5).mean()
    ma10 = closes_s.rolling(10).mean()
    ma20 = closes_s.rolling(20).mean()

    last = -1
    return {
        "symbol": symbol, "trade_date": trade_date,
        "close": round(float(closes_s.iloc[last]), 2),
        "ma5": round(float(ma5.iloc[last]), 2) if not pd.isna(ma5.iloc[last]) else None,
        "ma10": round(float(ma10.iloc[last]), 2) if not pd.isna(ma10.iloc[last]) else None,
        "ma20": round(float(ma20.iloc[last]), 2) if not pd.isna(ma20.iloc[last]) else None,
        "macd_dif": round(float(dif.iloc[last]), 4),
        "macd_dea": round(float(dea.iloc[last]), 4),
        "macd_bar": round(float(macd_bar.iloc[last]), 4),
        "macd_status": "金叉" if dif.iloc[last] > dea.iloc[last] else "死叉",
        "rsi_6": round(float(rsi.iloc[last]), 2),
        "kdj_k": None, "kdj_d": None, "kdj_j": None,  # KDJ 计算较复杂，暂略
    }


@router.get("/{task_id}/sandbox/realtime-indicators/{symbol}")
async def get_sandbox_realtime_indicators(task_id: str, symbol: str,
                                          trade_date: str = Query(None),
                                          phase_time: str = Query(None)):
    """回测模式下的"盘中估算"技术指标，基于本地分钟级快照 + 前日日线计算
    对应 Pi 的 get_realtime_indicators 工具回测路由
    16:00 前用分钟快照模拟盘中; 16:00 后用当天收盘 K 线
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta, datetime
    import pandas as pd

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"error": "无法确定回测日期"}

    trade_dt = dt_date.fromisoformat(trade_date)
    include_today = _is_day_finalized(phase_time)

    # ── 当前价 / 高 / 低 / 开盘 ──
    if phase_time and not include_today:
        # 16:00 前: 用分钟快照模拟盘中
        hh, mm = map(int, phase_time.split(":")[:2])
        minute_q = local_data.get_minute_quote(symbol, trade_dt, hh, mm)
        if not minute_q:
            return {"symbol": symbol, "error": f"无 {trade_date} {phase_time} 分钟快照"}
        current_price = minute_q["close"]
        # ── 累计日内高低点 (真正的日内分位, 而非单根分钟K线) ──
        cumulative = local_data.get_cumulative_intraday_high_low(
            symbol, trade_dt, hh, mm
        )
        if cumulative:
            day_high = cumulative["day_high"]
            day_low = cumulative["day_low"]
        else:
            day_high = minute_q["high"]
            day_low = minute_q["low"]
        day_open = minute_q["open"]
        prev_close = minute_q.get("pre_close", 0.0)
    else:
        # 16:00 后或无 phase_time: 用当天日 K（仅 16:00 后允许）
        if not include_today:
            return {"symbol": symbol, "error": "16:00 之前无完整日线，需提供 phase_time 取分钟快照"}
        day_q = local_data.get_daily_quote(symbol, trade_dt)
        if not day_q:
            return {"symbol": symbol, "error": "无行情数据"}
        current_price = day_q["close"]
        day_high = day_q["high"]
        day_low = day_q["low"]
        day_open = day_q["open"]
        prev_close = day_q.get("pre_close", day_q["close"])

    # ── 计算 MA5/MA10/MA20 ──
    # 取最近 20 个交易日收盘价（含 pre_close）
    closes = [prev_close]
    current_d = trade_dt - timedelta(days=1)
    while current_d >= trade_dt - timedelta(days=40) and len(closes) < 21:
        if current_d.weekday() < 5:
            q = local_data.get_daily_quote(symbol, current_d)
            if q:
                closes.insert(0, q["close"])
        current_d -= timedelta(days=1)
    closes_series = pd.Series(closes)

    ma5 = closes_series.rolling(5).mean().iloc[-1] if len(closes) >= 5 else closes_series.mean()
    ma10 = closes_series.rolling(10).mean().iloc[-1] if len(closes) >= 10 else ma5
    ma20 = closes_series.rolling(20).mean().iloc[-1] if len(closes) >= 20 else ma10

    # ── 计算 MACD ──
    ema12 = closes_series.ewm(span=12, adjust=False).mean()
    ema26 = closes_series.ewm(span=26, adjust=False).mean()
    dif = ema12 - ema26
    dea = dif.ewm(span=9, adjust=False).mean()
    macd_bar = 2 * (dif - dea)
    macd_dif_v = float(dif.iloc[-1])
    macd_dea_v = float(dea.iloc[-1])
    macd_bar_v = float(macd_bar.iloc[-1])

    # ── 计算 RSI(6) ──
    delta = closes_series.diff()
    gain = delta.clip(lower=0)
    loss = (-delta).clip(lower=0)
    avg_gain = gain.ewm(alpha=1/6, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1/6, adjust=False).mean()
    rs = avg_gain / avg_loss.replace(0, 1e-9)
    rsi6 = float((100 - 100 / (1 + rs)).iloc[-1])

    return {
        "symbol": symbol,
        "trade_date": trade_date,
        "phase_time": phase_time,
        "data_source": "intraday_estimate" if not include_today else "daily_confirmed",
        "current_price": round(current_price, 2),
        "high": round(day_high, 2),
        "low": round(day_low, 2),
        "open": round(day_open, 2),
        "prev_close": round(prev_close, 2),
        "ma5": round(float(ma5), 2),
        "ma10": round(float(ma10), 2),
        "ma20": round(float(ma20), 2),
        "macd_dif": round(macd_dif_v, 4),
        "macd_dea": round(macd_dea_v, 4),
        "macd_bar": round(macd_bar_v, 4),
        "rsi_6": round(rsi6, 2),
    }


@router.post("/{task_id}/sandbox/check-entry-filters")
async def check_entry_filters_sandbox(task_id: str, req: dict):
    """回测模式下的入场过滤检查（基于本地数据，不调实时接口）
    Pi 的 check_entry_filters 工具回测路由到此

    简化策略:
    - Layer1 技术面: 基于本地日线 + 分钟快照算 MA/MACD/RSI
    - Layer2 主力行为: 基于本地 moneyflow 算 5/10 日累计
    - Layer3 超买: RSI6 / KDJ-J
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta
    import pandas as pd

    symbol = req.get("symbol", "")
    phase_time = req.get("phase_time", "09:35")
    sector_net_inflow = req.get("sector_net_inflow", 0)
    volume_ratio = req.get("volume_ratio")  # 可选: 由调用方传入或端点估算

    # 获取任务 trade_date
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task or not task.current_day:
            return {"symbol": symbol, "error": "任务无当前日期"}
        trade_dt = task.current_day
    finally:
        db.close()

    # ── 复用 sandbox/realtime-indicators 函数（不通过 HTTP）──
    rt = await get_sandbox_realtime_indicators(
        task_id=task_id, symbol=symbol,
        trade_date=trade_dt.isoformat(),
        phase_time=phase_time,
    )
    if isinstance(rt, dict) and rt.get("error"):
        return {"symbol": symbol, "error": rt["error"]}

    # ── 量比估算: 调用方未传时，从分钟快照的换手率反推（与实时端点公式一致） ──
    if volume_ratio is None and phase_time and rt.get("data_source") == "intraday_estimate":
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            mq = local_data.get_minute_quote(symbol, trade_dt, hh, mm)
            tr = mq.get("turnover_rate", 0) if mq else 0
            if tr > 0:
                volume_ratio = round(tr / 2.0, 2)  # 历史日均换手率约 2%
        except Exception:
            volume_ratio = None

    # ── Layer 1: 技术面 ──
    ma5 = rt.get("ma5", 0)
    ma10 = rt.get("ma10", 0)
    ma20 = rt.get("ma20", 0)
    current_price = rt.get("current_price", 0)
    macd_dif = rt.get("macd_dif", 0)
    macd_dea = rt.get("macd_dea", 0)
    macd_bar = rt.get("macd_bar", 0)
    rsi_6 = rt.get("rsi_6", 50)

    above_ma5 = current_price > ma5
    above_ma20 = current_price > ma20
    macd_golden = macd_dif > macd_dea
    ma5_gt_ma20 = ma5 > ma20

    layer1_pass = above_ma5 and above_ma20 and macd_golden
    if not ma5_gt_ma20 and sector_net_inflow > 0:
        # MA5<MA20 时用板块资金流兜底
        layer1_pass = sector_net_inflow > 0

    # ── Layer 2: 主力行为 (用本地 moneyflow 5 个交易日累计, 含当日 B2 缩放) ──
    # 修复: 
    #   1) 原 cursor = trade_dt - 1day → 跳过当日, 导致突破日的方向信号被忽略
    #      → 现在从当天开始, 用 B2 缩放当天值 (反未来函数: 盘中期不会拿到 EOD 全量)
    #   2) Tushare "主力" = 超大单净额 + 大单净额 = (buy_elg - sell_elg) + (buy_lg - sell_lg)
    try:
        main_net_5d = 0.0
        days_collected = 0
        cursor = trade_dt  # ← 从今天开始, 不是昨天
        while days_collected < 5 and cursor.year > trade_dt.year - 5:
            if cursor.weekday() < 5:
                mf = local_data.get_moneyflow(symbol, cursor)
                if mf:
                    lg_net = float(mf.get("buy_elg_amount", 0) or 0) - float(mf.get("sell_elg_amount", 0) or 0)
                    md_net = float(mf.get("buy_lg_amount", 0) or 0) - float(mf.get("sell_lg_amount", 0) or 0)
                    day_val = lg_net + md_net
                    # 当天记录: 应用 B2 缩放 (反未来函数, 盘中期不拿全天 EOD)
                    if cursor == trade_dt and phase_time:
                        try:
                            hh, mm = map(int, phase_time.split(":")[:2])
                            wi = local_data.get_moneyflow_intraday_weight(symbol, trade_dt, hh, mm)
                            if wi:
                                day_val *= wi["weight"]
                        except Exception:
                            pass  # 缩放失败用原值(保守: 可能略高估但不会漏信号)
                    main_net_5d += day_val
                    days_collected += 1
            cursor -= timedelta(days=1)
        if days_collected < 5:
            print(f"[check_entry_filters] 5日主力只累计了 {days_collected} 天 (可能早期股票无数据)", flush=True)
    except Exception:
        main_net_5d = 0
    layer2_pass = main_net_5d > 0
    # 转换为 亿 供前端显示 (parquet 原始单位是 元, tools.ts 显示后缀为"亿")
    main_net_5d_e8 = round(main_net_5d / 1e8, 4)

    # ── Layer 3: 超买 ──
    rsi_overbought = rsi_6 > 70
    layer3_pass = not rsi_overbought

    # ── 决策 ──
    all_pass = layer1_pass and layer2_pass and layer3_pass
    symbol_name = symbol.split(".")[-1] if "." in symbol else symbol

    # ── 涨幅 + 买入确认（对齐实时端点 SOP 规则）──
    prev_close = rt.get("prev_close", 0) or 0
    if prev_close > 0:
        change_pct = round((current_price - prev_close) / prev_close * 100, 2)
    else:
        change_pct = 0.0

    if change_pct <= 2:
        bc_action = "可建仓"
        bc_wait = 0
        bc_allow = all_pass
        bc_ratio = 1.0 if all_pass else 0.5
    elif change_pct <= 8:
        bc_action = "等2-3分钟，量比>1.5才入场"
        bc_wait = 2
        vol_ok = (volume_ratio is not None and volume_ratio > 1.5)
        bc_allow = all_pass and vol_ok
        bc_ratio = 0.5 if (all_pass and vol_ok) else 0.0
    else:
        bc_action = "放弃（涨幅>8%，不追涨）"
        bc_wait = 0
        bc_allow = False
        bc_ratio = 0.0

    # ── 日内分位 (累计日内高低点, 非单根分钟K线) ──
    intraday_percentile = None
    if phase_time and current_price > 0:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            cum = local_data.get_cumulative_intraday_high_low(
                symbol, trade_dt, hh, mm
            )
            if cum:
                intraday_percentile = cum["intraday_percentile"]
        except Exception:
            pass

    return {
        "symbol": symbol,
        "name": symbol_name,
        "trade_date": trade_dt.isoformat(),
        "data_source": "backtest_local",
        "current_price": current_price,
        "change_pct": change_pct,
        "ma5": ma5, "ma10": ma10, "ma20": ma20,
        "macd_dif": macd_dif, "macd_dea": macd_dea, "macd_bar": macd_bar,
        "rsi_6": rsi_6,
        "main_net_5d": main_net_5d_e8,      # 亿 (tools.ts 显示为 `亿`)
        "volume_ratio": volume_ratio,
        "intraday_percentile": intraday_percentile,
        "layer1_tech": {"pass": layer1_pass, "above_ma5": above_ma5, "above_ma20": above_ma20, "macd_golden": macd_golden, "ma5_gt_ma20": ma5_gt_ma20},
        "layer2_capital": {"pass": layer2_pass, "main_net_5d": main_net_5d_e8},  # 亿
        "layer3_overbought": {"pass": layer3_pass, "rsi_6": rsi_6, "rsi_overbought": rsi_overbought},
        "tech": {
            "pass": all_pass, "summary": "✅ 全部通过" if all_pass else "⚠️ 部分过滤未通过",
            "intraday_percentile": intraday_percentile,
        },
        "buy_confirmation": {
            "allow": bc_allow, "ratio": bc_ratio,
            "action": bc_action, "wait_minutes": bc_wait,
            "volume_ratio": volume_ratio,
            "volume_ratio_ok": (volume_ratio is not None and volume_ratio > 1.5) if 2 < change_pct <= 8 else None,
        },
    }


@router.get("/{task_id}/sandbox/moneyflow/{symbol}")
async def get_sandbox_moneyflow(task_id: str, symbol: str,
                                 trade_date: str = Query(None),
                                 limit: int = Query(5, ge=1, le=30),
                                 phase_time: str = Query(None)):
    """返回个股历史资金流向（从本地 moneyflow.parquet），Pi 的 get_moneyflow 回测路由到此

    反未来函数:
      - moneyflow.parquet 是日频(无盘中累计), 盘中"当日"用 B2 成交额加权缩放
      - 缩放公式: scale = stock_1min 累计成交额 / 全天总成交额 (下限 0.15)
      - 09:30 之前 (盘前) → scale=0, 当日字段 = 0 (尚未开市)
      - 15:00+ (盘后) → scale=1.0, 当日字段 = 全天日终实际值 (EOD 落盘)
      - 5/10 日累计: 取 5/10 个历史交易日, 含当日缩放部分
    字段:
      - records: 历史资金流向(每条 1 日), 按日期倒序
      - today:  当日资金流汇总 (含 scale 字段, 供 Pi 显示估算比例)
      - cum5/cum10: 5/10 日累计 (基于日终值, 不缩放)
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"records": [], "error": "无法确定回测日期"}

    end_dt = dt_date.fromisoformat(trade_date)
    start_dt = end_dt - timedelta(days=max(limit * 3, 30))  # 拉宽些保证 5/10 日累计有数据

    # ── 计算缩放系数 ──
    # 优先 B2 (成交额加权), 回退 B1 (时间线性), 盘前=0, 盘后=1.0
    scale = 1.0  # 缺省按盘后处理
    scale_basis = "time_linear"  # 默认
    scale_debug = {}
    phase_minutes = None
    is_intraday = False
    is_pre_market = False
    if phase_time:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            cur = hh * 60 + mm
            phase_minutes = cur
            market_open = 9 * 60 + 30   # 09:30
            market_close = 15 * 60      # 15:00
            finalized = 16 * 60          # 16:00 视为 EOD 落盘
            if cur < market_open:
                scale = 0.0
                is_pre_market = True
                scale_basis = "pre_market_zero"
            elif cur < market_close:
                is_intraday = True
                # B2 尝试: 用 stock_1min 成交额占比作为权重
                weight_info = local_data.get_moneyflow_intraday_weight(symbol, end_dt, hh, mm)
                if weight_info is not None:
                    scale = weight_info["weight"]
                    scale_basis = "amount_weighted"
                    scale_debug = {
                        "target_amount": weight_info["target_amount"],
                        "total_amount": weight_info["total_amount"],
                    }
                    # 分钟K线方向修正: 用蜡烛图形态估算实际买卖倾向
                    bias_info = local_data.get_minute_flow_bias(symbol, end_dt, hh, mm)
                    if bias_info:
                        bias = bias_info["bias"]
                        scale = round(scale * (1 + bias * 0.3), 6)
                        scale_basis = f"amount_weighted_candle_adj"
                        scale_debug["candle_bias"] = bias
                        scale_debug["candle_count"] = bias_info["candle_count"]
                else:
                    # B1 回退: 时间均匀
                    scale = round((cur - market_open) / (market_close - market_open), 4)
                    scale_basis = "time_linear_fallback"
            else:
                # 15:00 之后: 全量
                scale = 1.0
                scale_basis = "post_market_eod"
        except Exception:
            scale = 1.0
            scale_basis = "error_fallback_eod"
    else:
        # phase_time 缺省: 保守按盘后(EOD)处理
        scale = 1.0
        scale_basis = "phase_missing_eod"

    # 历史记录从 end_dt 倒推 (不再排除当日)
    current = end_dt
    raw_records = []
    while current >= start_dt and len(raw_records) < max(limit, 11):
        mf = local_data.get_moneyflow(symbol, current)
        if mf:
            raw_records.append(mf)
        current -= timedelta(days=1)

    # ── 计算"主力"和累计 (Tushare moneyflow 原生不含"主力"字段) ──
    # 主力 = 超大单 + 大单 (买卖差额), 与东财/同花顺"主力"口径一致
    def _with_main(mf: dict, apply_scale: bool = False, scale_factor: float = 1.0) -> dict:
        buy_main = float(mf.get("buy_elg_amount", 0) or 0) + float(mf.get("buy_lg_amount", 0) or 0)
        sell_main = float(mf.get("sell_elg_amount", 0) or 0) + float(mf.get("sell_lg_amount", 0) or 0)
        out = dict(mf)
        if apply_scale:
            out["buy_elg_amount"]  = round(out.get("buy_elg_amount", 0) * scale_factor, 2)
            out["sell_elg_amount"] = round(out.get("sell_elg_amount", 0) * scale_factor, 2)
            out["buy_lg_amount"]   = round(out.get("buy_lg_amount", 0) * scale_factor, 2)
            out["sell_lg_amount"]  = round(out.get("sell_lg_amount", 0) * scale_factor, 2)
            out["net_mf_amount"]   = round(out.get("net_mf_amount", 0) * scale_factor, 2)
            out["net_mf_vol"]      = round(out.get("net_mf_vol", 0) * scale_factor, 2)
            buy_main *= scale_factor
            sell_main *= scale_factor
        out["main_net_amount"] = round(buy_main - sell_main, 2)  # 主力净流入额(元)
        out["main_buy_amount"] = round(buy_main, 2)              # 主力买入额
        out["main_sell_amount"] = round(sell_main, 2)            # 主力卖出额
        out["scaled"] = apply_scale
        out["scale"] = scale_factor
        return out

    # 当日 (records[0]) 按 scale 缩放, 历史日 (records[1:]) 用全量
    records = []
    for i, r in enumerate(raw_records):
        is_today_row = (i == 0)
        records.append(_with_main(r, apply_scale=is_today_row, scale_factor=scale if is_today_row else 1.0))

    today = records[0] if records else None

    # 5/10 日累计: 注意 raw_records 顺序已是"由新到旧"
    def _sum_main(rows):
        return {
            "window_days": len(rows),
            "main_net_sum": round(sum(r["main_net_amount"] for r in rows), 2),
            "net_mf_sum":   round(sum(r["net_mf_amount"]  for r in rows), 2),
            "start_date":   rows[-1]["date"] if rows else None,
            "end_date":     rows[0]["date"]  if rows else None,
        }

    cum5  = _sum_main(records[:5])  if len(records) >= 5  else _sum_main(records)
    cum10 = _sum_main(records[:10]) if len(records) >= 10 else _sum_main(records)

    # caveat: 明确告知 Pi 这是"估算的盘中累计"
    if is_pre_market:
        data_freshness = "pre_market_zero"
        caveat = f"⚠️ 盘前 phase_time={phase_time}, 资金流尚未开始, 全部字段=0"
    elif is_intraday:
        data_freshness = "intraday_eod_scaled"
        # caveat 仅供调试日志, 不暴露内部缩放参数给 AI
        if scale_basis in ("amount_weighted", "amount_weighted_candle_adj"):
            caveat = "盘中估算 (基于成交额进度, 非EOD全量)"
        elif scale_basis == "time_linear_fallback":
            caveat = "盘中估算 (时间线性缩放, 误差可能较大)"
        else:
            caveat = "盘中估算"
    else:
        data_freshness = "post_market_eod"
        caveat = None

    return {
        "symbol": symbol,
        "trade_date": trade_date,
        "phase_time": phase_time,
        "phase_minutes": phase_minutes,
        "is_intraday": is_intraday,
        "is_pre_market": is_pre_market,
        "data_freshness": data_freshness,
        "scale": scale,
        "scale_basis": scale_basis,
        "scale_debug": scale_debug,
        "caveat": caveat,
        "today": today,
        "cum5": cum5,
        "cum10": cum10,
        "records": records[:limit],  # 受 limit 截断, 与旧契约兼容
        "count": len(records[:limit]),
        "total_available": len(records),
    }


@router.get("/{task_id}/sandbox/concept-fund-flow")
async def get_sandbox_concept_fund_flow(task_id: str,
                                          trade_date: str = Query(None),
                                          limit: int = Query(15, ge=1, le=50),
                                          sort_by: str = Query("main_net"),
                                          phase_time: str = Query(None, description="Pi 调用时刻 HH:MM，盘中期(null/<16:00)用前日数据")):
    """回测模式下的概念板块资金流（基于本地 parquet）
    对应 Pi 的 get_concept_fund_flow 工具回测路由
    反未来函数: 9:30 后准入当日渐进数据 (B2 成交额加权缩放), 盘前用前日数据
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"sectors": [], "error": "无法确定回测日期"}

    end_dt = dt_date.fromisoformat(trade_date)
    include_today = _is_moneyflow_available(phase_time)
    base_dt = end_dt if include_today else (end_dt - timedelta(days=1))
    while base_dt.weekday() >= 5:
        base_dt -= timedelta(days=1)

    sectors = local_data.get_concept_flow(base_dt, top_n=limit)

    # ── 盘中 pct_change 修正 (避免"全天排名冻结" + 反未来函数) ──
    # 问题: parquet 的 pct_change 是当日 EOD 值, 盘中不变 → 排名永远相同 → Pi 误判"市场冻结"
    # 修复: 用 lead_stock 的分钟行情估算板块当日实时涨跌
    sector_today_pct = {}
    if include_today and phase_time and sectors:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            for s in sectors:
                ls = s.get("lead_stock", "")
                if ls:
                    mq = local_data.get_minute_quote(ls, end_dt, hh, mm)
                    if mq:
                        cur = float(mq.get("close", 0))
                        prev = float(mq.get("last_close", 0)) or float(mq.get("pre_close", 0))
                        if cur > 0 and prev > 0:
                            sector_today_pct[ls] = round((cur / prev - 1) * 100, 2)
            # 对有 lead_stock 的 sector, 覆盖 pct_change 为盘中实时值
            if sector_today_pct:
                for s in sectors:
                    ls = s.get("lead_stock", "")
                    if ls in sector_today_pct:
                        s["pct_change"] = sector_today_pct[ls]
        except Exception:
            pass  # 静默降级: 用原 EOD pct_change

    if sort_by == "pct_change":
        sectors = sorted(sectors, key=lambda x: x.get("pct_change", 0), reverse=True)

    # ── 方案B: 成分股分钟K线估算盘中资金流 (替代 B2 EOD 缩放) ──
    # 原理: 不用 EOD net_amount × 权重, 而是从板块成分股的真实分钟蜡烛图
    #       估算买卖方向和净额。这消除了 B2 的 "EOD 排名保留" 偏差,
    #       09:35 的板块排名可以完全不同于 EOD, 更接近真实市场的不确定性。
    scale = 1.0
    scale_basis = "minute_component_stocks"
    scale_debug: dict = {}
    if include_today and phase_time and sectors:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            for s in sectors:
                sector_name = s.get("name", "")
                if not sector_name:
                    continue
                # 方案B: 从成分股分钟K线估算板块资金流
                ib = local_data.get_sector_intraday_bias(
                    sector_name, end_dt, hh, mm,
                    sector_type="concept", component_limit=8,
                )
                if ib and ib["sample_count"] > 0:
                    s["net_amount"] = ib["estimated_net"]
                    s["intraday_bias"] = ib["bias"]          # 买卖倾向
                    s["sample_stocks"] = ib["sample_count"]   # 有效成分股数
                    s["detail"] = ib.get("detail", [])[:3]    # 前3详情
                else:
                    # 降级: 用 lead_stock 的分钟 bias 估算
                    ls = s.get("lead_stock", "")
                    if ls:
                        bias_info = local_data.get_minute_flow_bias(ls, end_dt, hh, mm)
                        if bias_info:
                            eod_net = abs(float(s.get("net_amount", 0)))
                            s["net_amount"] = round(eod_net * bias_info["bias"] * 0.15, 4)
                            s["intraday_bias"] = bias_info["bias"]
                            s["sample_stocks"] = 1
                        else:
                            s["net_amount"] = 0.0
                    else:
                        s["net_amount"] = 0.0
            scale_basis = "minute_component_stocks"
        except Exception:
            scale = 1.0
            scale_basis = "eod_full"

    if include_today:
        if scale_basis == "eod_full":
            data_freshness = "today_eod"
            caveat = None
        elif scale_basis == "pre_market_zero":
            data_freshness = "pre_market_zero"
            caveat = f"⚠️ 盘前 phase_time={phase_time}, 资金流尚未开始, 全部 net_amount=0"
        else:
            data_freshness = "intraday_component_estimated"
            # 不暴露内部算法细节给 AI
            caveat = None
    else:
        data_freshness = "yesterday_eod"
        caveat = f"⚠️ 资金流数据为昨日({base_dt.isoformat()})日终, 不是今日盘中实时"

    return {
        "sectors": sectors, "trade_date": base_dt.isoformat(),
        "source_trade_date": end_dt.isoformat(),
        "data_source": "backtest_local",
        "sort_by": sort_by,
        "data_freshness": data_freshness,
        "scale": scale,
        "scale_basis": scale_basis,
        "scale_debug": scale_debug,
        "phase_time": phase_time,
        "is_intraday": include_today and phase_time is not None and scale_basis not in ("eod_full", "pre_market_zero"),
        "caveat": caveat,
    }


@router.get("/{task_id}/sandbox/industry-fund-flow")
async def get_sandbox_industry_fund_flow(task_id: str,
                                          trade_date: str = Query(None),
                                          limit: int = Query(15, ge=1, le=50),
                                          sort_by: str = Query("main_net"),
                                          phase_time: str = Query(None, description="Pi 调用时刻 HH:MM，盘中期(null/<16:00)用前日数据")):
    """回测模式下的行业板块资金流（基于本地 parquet）
    反未来函数: 用 phase_time 判断; 9:30 后准入当日渐进数据 (B2 成交额加权缩放), 盘前用前日数据
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"sectors": [], "error": "无法确定回测日期"}

    end_dt = dt_date.fromisoformat(trade_date)
    include_today = _is_moneyflow_available(phase_time)
    base_dt = end_dt if include_today else (end_dt - timedelta(days=1))
    while base_dt.weekday() >= 5:
        base_dt -= timedelta(days=1)

    sectors = local_data.get_industry_flow(base_dt, top_n=limit)

    # ── 盘中 pct_change 修正 (同概念端点: 用 lead_stock 分钟行情) ──
    if include_today and phase_time and sectors:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            for s in sectors:
                ls = s.get("lead_stock", "") or s.get("lead_stock_code", "")
                if ls:
                    mq = local_data.get_minute_quote(ls, end_dt, hh, mm)
                    if mq:
                        cur = float(mq.get("close", 0))
                        prev = float(mq.get("last_close", 0)) or float(mq.get("pre_close", 0))
                        if cur > 0 and prev > 0:
                            s["pct_change"] = round((cur / prev - 1) * 100, 2)
        except Exception:
            pass

    if sort_by == "pct_change":
        sectors = sorted(sectors, key=lambda x: x.get("pct_change", 0), reverse=True)

    # ── 方案B: 成分股分钟K线估算盘中资金流 (替代 B2 EOD 缩放) ──
    scale = 1.0
    scale_basis = "minute_component_stocks"
    scale_debug: dict = {}
    if include_today and phase_time and sectors:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            for s in sectors:
                sector_name = s.get("name", "")
                if not sector_name:
                    continue
                ib = local_data.get_sector_intraday_bias(
                    sector_name, end_dt, hh, mm,
                    sector_type="industry", component_limit=8,
                )
                if ib and ib["sample_count"] > 0:
                    s["net_amount"] = ib["estimated_net"]
                    s["intraday_bias"] = ib["bias"]
                    s["sample_stocks"] = ib["sample_count"]
                    s["detail"] = ib.get("detail", [])[:3]
                else:
                    ls = s.get("lead_stock", "") or s.get("lead_stock_code", "")
                    if ls:
                        bias_info = local_data.get_minute_flow_bias(ls, end_dt, hh, mm)
                        if bias_info:
                            eod_net = abs(float(s.get("net_amount", 0)))
                            s["net_amount"] = round(eod_net * bias_info["bias"] * 0.15, 4)
                            s["intraday_bias"] = bias_info["bias"]
                            s["sample_stocks"] = 1
                        else:
                            s["net_amount"] = 0.0
                    else:
                        s["net_amount"] = 0.0
            scale_basis = "minute_component_stocks"
        except Exception:
            scale = 1.0
            scale_basis = "eod_full"

    if include_today:
        if scale_basis == "eod_full":
            data_freshness = "today_eod"
            caveat = None
        elif scale_basis == "pre_market_zero":
            data_freshness = "pre_market_zero"
            caveat = f"⚠️ 盘前 phase_time={phase_time}, 资金流尚未开始, 全部 net_amount=0"
        else:
            data_freshness = "intraday_component_estimated"
            caveat = None
    else:
        data_freshness = "yesterday_eod"
        caveat = f"⚠️ 行业资金流数据为昨日({base_dt.isoformat()})日终, 不是今日盘中实时"

    return {
        "sectors": sectors, "trade_date": base_dt.isoformat(),
        "source_trade_date": end_dt.isoformat(),
        "data_source": "backtest_local",
        "sort_by": sort_by,
        "data_freshness": data_freshness,
        "scale": scale,
        "scale_basis": scale_basis,
        "scale_debug": scale_debug,
        "phase_time": phase_time,
        "is_intraday": include_today and phase_time is not None and scale_basis not in ("eod_full", "pre_market_zero"),
        "caveat": caveat,
    }


@router.get("/{task_id}/sandbox/market-moneyflow")
async def get_sandbox_market_moneyflow(task_id: str,
                                        trade_date: str = Query(None),
                                        phase_time: str = Query(None, description="Pi 调用时刻 HH:MM，盘中期(null/<16:00)用前日数据")):
    """回测模式下的大盘资金流向（基于本地 parquet）
    反未来函数: 用 phase_time 判断; 9:30 后准入当日渐进数据 (B2 成交额加权缩放), 盘前用前日数据
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"data": None, "error": "无法确定回测日期"}

    end_dt = dt_date.fromisoformat(trade_date)
    include_today = _is_moneyflow_available(phase_time)

    # B2: 盘中按 get_market_moneyflow_intraday 缩放
    if include_today and phase_time:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
            mi = local_data.get_market_moneyflow_intraday(end_dt, hh, mm)
            if mi is not None:
                data_freshness = "intraday_eod_scaled" if mi["basis"] not in ("eod_full", "pre_market_zero") else "today_eod"
                basis = mi["basis"]
                # caveat 仅供调试日志, 不暴露内部缩放参数给 AI
                caveat = "盘中估算 (基于全市场加总缩放)" if mi["weight"] != 1.0 else None
                return {
                    "data": {
                        "trade_date": end_dt.isoformat(),
                        "data_source": "backtest_local_intraday_scaled",
                        "net_amount": round(mi["intraday_net"] / 1e8, 4),
                        "buy_elg": round(mi["intraday_buy_elg"] / 1e8, 4),
                        "buy_lg": round(mi["intraday_buy_lg"] / 1e8, 4),
                        "buy_md": round(mi["intraday_buy_md"] / 1e8, 4),
                        "buy_sm": round(mi["intraday_buy_sm"] / 1e8, 4),
                        "main_net": round(mi["intraday_main_net"] / 1e8, 4),
                        "eod_net_amount": round(mi["eod_net_amount"] / 1e8, 4),
                        "scale": mi["weight"],
                        "scale_basis": basis,
                    },
                    "source_trade_date": end_dt.isoformat(),
                    "data_freshness": data_freshness,
                    "phase_time": phase_time,
                    "is_intraday": basis not in ("eod_full", "pre_market_zero"),
                    "caveat": caveat,
                }
        except Exception:
            pass  # 走 fallback 分支

    # 盘前 / 盘后 / phase_time 缺失: 用前日(或当日 EOD)原数据
    base_dt = end_dt if include_today else (end_dt - timedelta(days=1))
    while base_dt.weekday() >= 5:
        base_dt -= timedelta(days=1)

    mf = local_data.get_market_flow(base_dt)
    if not mf:
        return {"data": None, "error": f"无 {base_dt} 大盘资金流数据"}

    return {
        "data": {
            **mf,
            "trade_date": base_dt.isoformat(),
            "data_source": "backtest_local",
        },
        "source_trade_date": end_dt.isoformat(),
        "data_freshness": "today_eod" if include_today else "yesterday_eod",
        "caveat": None if include_today else f"⚠️ 大盘资金流数据为昨日({base_dt.isoformat()})日终, 不是今日盘中实时",
    }


@router.get("/{task_id}/sandbox/realtime-sector-pct")
async def get_sandbox_realtime_sector_pct(task_id: str,
                                            trade_date: str = Query(None),
                                            phase_time: str = Query(None, description="HH:MM 模拟时刻"),
                                            theme_top_n: int = Query(15, ge=1, le=50)):
    """回测模式下的盘中实时行业/主题涨跌幅（反未来函数核心数据源）

    数据源: data/backtest/指数数据/index_1min/
    - 10 个中证一级行业指数 (000032~000041.SH) — 行业大类盘中涨跌幅
    - 287 个主题指数 — 题材/概念盘中涨跌幅 (按涨幅排序 top N)

    用途: 解决回测里 Pi "永远看后视镜"的问题, 盘中 phase (09:35/10:35/13:35/14:30)
          能拿到真实的"截至当前 phase_time 的行业/主题盘中累计涨跌"
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"error": "无法确定回测日期", "industries": [], "themes": []}

    if not phase_time:
        return {"error": "需提供 phase_time=HH:MM", "industries": [], "themes": []}

    try:
        hh, mm = map(int, phase_time.split(":")[:2])
    except Exception:
        return {"error": f"phase_time 格式错误: {phase_time}", "industries": [], "themes": []}

    trade_dt = dt_date.fromisoformat(trade_date)
    result = local_data.get_realtime_sector_pct(trade_dt, hh, mm, theme_top_n=theme_top_n)
    return result


@router.get("/{task_id}/sandbox/concept-mapping")
async def get_sandbox_concept_mapping(task_id: str,
                                       trade_date: str = Query(None),
                                       concept_name: str = Query(None),
                                       symbol: str = Query(None),
                                       limit: int = Query(100, ge=1, le=500)):
    """回测模式下的概念成分股（按 trade_date 拉取当日成分）
    数据源: Tushare dc_index(trade_date) → dc_member(trade_date, ts_code | con_code)

    三种模式:
    - 仅 trade_date: 返回 trade_date 当天有效概念列表（按涨跌幅排序）
    - trade_date + concept_name: 返回该概念在 trade_date 当天的成分股
    - trade_date + symbol: 反查该股票所属的概念板块（持仓归因场景）

    ⚠️ 关键: 必须用 dc_index/dc_member（东财，BKxxxx.DC），
        不能用 ths_index/ths_member（同花顺，885xxx.TI）—— 两者代码体系不同
    """
    from app.config import get_settings
    from datetime import date as dt_date

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"concepts": [], "stocks": [], "error": "无法确定回测日期"}

    settings = get_settings()
    token = settings.get_tushare_token()
    if not token:
        return {"concepts": [], "stocks": [], "error": "Tushare token 未配置"}

    try:
        import tushare as ts
        pro = ts.pro_api(token)
        td = dt_date.fromisoformat(trade_date)
        td_str = td.strftime("%Y%m%d")

        # ── 模式 3: 按股票代码反查所属概念（持仓归因）──
        if symbol and not concept_name:
            # 归一化 symbol 为 ts_code 格式 (SH600519 → 600519.SH; 002156 → 002156.SZ)
            s = symbol.strip().upper()
            if s.startswith(("SH", "SZ")) and "." not in s:
                s = s[2:] + "." + s[:2]
            elif "." not in s and len(s) == 6:
                s = s + (".SH" if s.startswith("6") else ".SZ")
            ts_code = s
            # dc_member(con_code) 拿股票所属所有概念（取 trade_date 当天）
            df_member = pro.dc_member(trade_date=td_str, con_code=ts_code)
            if df_member is None or df_member.empty:
                return {"symbol": ts_code, "trade_date": trade_date,
                        "concepts": [], "concept_count": 0,
                        "warning": f"{ts_code} 在 {trade_date} 无概念归属数据"}

            concepts = []
            for _, r in df_member.iterrows():
                concepts.append({
                    "ts_code": str(r.get("ts_code", "")),
                    "concept_name": str(r.get("name", "") or ""),  # name 在这里指概念名
                })
            return {
                "symbol": ts_code,
                "trade_date": trade_date,
                "concept_count": len(concepts),
                "concepts": concepts[:limit],
                "data_source": "tushare_dc_member_con_code",
            }

        if not concept_name:
            # 列出 trade_date 当天有效概念（按涨跌幅倒序）
            df_idx = pro.dc_index(trade_date=td_str)
            if df_idx is None or df_idx.empty:
                # 降级: 往前找最近一个交易日
                from datetime import timedelta
                for back in range(1, 10):
                    prev_str = (td - timedelta(days=back)).strftime("%Y%m%d")
                    df_idx = pro.dc_index(trade_date=prev_str)
                    if df_idx is not None and not df_idx.empty:
                        td_str = prev_str
                        break
                if df_idx is None or df_idx.empty:
                    return {"concepts": [], "stocks": [], "total": 0, "error": "dc_index 9 日内均无数据"}

            # 按涨跌幅排序
            if "pct_change" in df_idx.columns:
                df_idx = df_idx.sort_values("pct_change", ascending=False)
            concepts = df_idx.head(limit).to_dict("records")
            return {
                "concepts": [
                    {
                        "sector_name": c.get("name", ""),
                        "ts_code": c.get("ts_code", ""),
                        "pct_change": float(c.get("pct_change", 0) or 0),
                        "leading": str(c.get("leading", "") or ""),
                        "leading_code": str(c.get("leading_code", "") or ""),
                        "up_num": int(c.get("up_num", 0) or 0),
                        "down_num": int(c.get("down_num", 0) or 0),
                    }
                    for c in concepts
                ],
                "total": len(concepts),
                "trade_date": td_str,
                "source_trade_date": trade_date,
                "data_source": "tushare_dc_index",
            }
        else:
            # Step 1: 中文名 → BKxxxx.DC (用 dc_index + name 模糊匹配)
            df_idx = pro.dc_index(trade_date=td_str, name=concept_name.strip())
            if df_idx is None or df_idx.empty:
                # 降级: 不限日期, 全量查
                df_idx_all = pro.dc_index()
                if df_idx_all is not None and not df_idx_all.empty:
                    matched_all = df_idx_all[df_idx_all["name"].str.contains(
                        concept_name.strip(), na=False, regex=False)]
                    if not matched_all.empty:
                        # 优先取最近日期的匹配
                        if "trade_date" in matched_all.columns:
                            matched_all = matched_all.sort_values("trade_date", ascending=False)
                        df_idx = matched_all.head(1)
                if df_idx is None or df_idx.empty:
                    return {"concept": concept_name, "stocks": [], "stock_count": 0,
                            "error": f"未找到概念 [{concept_name}]"}

            con_code = df_idx.iloc[0]["ts_code"]
            con_name = df_idx.iloc[0].get("name", concept_name)

            # Step 2: dc_member(trade_date, ts_code) 拿当天成分
            df_member = pro.dc_member(trade_date=td_str, ts_code=con_code)
            if df_member is None or df_member.empty:
                return {
                    "concept": con_name, "ts_code": con_code,
                    "trade_date": trade_date, "stocks": [], "stock_count": 0,
                    "warning": f"{con_code} 在 {trade_date} 无成分数据（可能非交易日或概念已下架）",
                }

            df_member = df_member.head(limit)
            return {
                "concept": con_name,
                "ts_code": con_code,
                "trade_date": trade_date,
                "stock_count": len(df_member),
                "stocks": [
                    {
                        "ts_code": str(r.get("con_code", "")),
                        "symbol": str(r.get("con_code", "")).split(".")[0] if r.get("con_code") else "",
                        "name": str(r.get("name", "") or ""),
                    }
                    for _, r in df_member.iterrows()
                ],
                "data_source": "tushare_dc_member",
            }
    except Exception as e:
        return {"concepts": [], "stocks": [], "error": f"Tushare 调用失败: {e}"}


@router.get("/{task_id}/sandbox/indices")
async def get_sandbox_indices(task_id: str,
                                trade_date: str = Query(None),
                                phase_time: str = Query(None, description="Pi 调用时刻 HH:MM（用于反查未来函数）")):
    """回测模式下的市场指数
    数据源: 本地 index_basic.parquet (指数列表) + Tushare index_daily (历史日线)
    支持的指数: 上证 000001.SH / 深证 399001.SZ / 沪深300 000300.SH / 创业板 399006.SZ / 科创50 000688.SH

    ⚠️ 反未来函数 (Tushare 官方明确: 大盘指数无历史分钟数据, 只能取日线字段):
       - 盘前 (phase_time < 09:30):  用 trade_date 前一日收盘, change_pct=0
       - 盘中 (09:30 <= phase_time < 16:00): 用 trade_date 当日 OPEN
          (9:35 时 current_price=9:30 开盘价, high/low=open 表示"今日已开盘 5 分钟但无指数分钟数据",
           change_pct=0 因为"今日还未收盘")
       - 盘后 (phase_time >= 16:00):  全量日线数据 (open/high/low/close + 真实 change_pct)
    """
    from app.config import get_settings
    from datetime import date as dt_date, timedelta
    from pathlib import Path
    import pandas as pd

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"indices": [], "error": "无法确定回测日期"}

    # 关注的 5 大指数
    target_codes = ["000001.SH", "399001.SZ", "000300.SH", "399006.SZ", "000688.SH"]
    name_map = {
        "000001.SH": "上证指数", "399001.SZ": "深证成指",
        "000300.SH": "沪深300", "399006.SZ": "创业板指", "000688.SH": "科创50",
    }

    indices = []
    data_source = "tushare_index_daily"

    # Step 1: 读本地 index_basic 拿指数名称
    index_basic_path = (Path(__file__).parent.parent.parent.parent / "data" / "backtest" / "指数数据" / "index_basic.parquet")
    if not index_basic_path.exists():
        return {"indices": [], "error": "index_basic.parquet 不存在"}

    try:
        df_basic = pd.read_parquet(index_basic_path)
        basic_map = {row["ts_code"]: row["name"] for _, row in df_basic.iterrows()}
    except Exception as e:
        return {"indices": [], "error": f"读 index_basic 失败: {e}"}

    # Step 2: Tushare index_daily 拿历史日线
    settings = get_settings()
    token = settings.get_tushare_token()
    if not token:
        return {"indices": [], "error": "Tushare token 未配置"}

    try:
        import tushare as ts
        pro = ts.pro_api(token)
        td = dt_date.fromisoformat(trade_date)
        td_str = td.strftime("%Y%m%d")

        # ── 反未来函数: 按 phase_time 决定取数策略 ──
        if phase_time:
            try:
                hh, mm = map(int, phase_time.split(":")[:2])
                cur_minutes = hh * 60 + mm
            except Exception:
                cur_minutes = None
        else:
            cur_minutes = None

        is_pre_market = cur_minutes is not None and cur_minutes < 9 * 60 + 30
        is_intraday = cur_minutes is not None and 9 * 60 + 30 <= cur_minutes < 16 * 60
        is_post_market = cur_minutes is not None and cur_minutes >= 16 * 60

        if is_pre_market:
            # 盘前: 用前一日收盘
            end_d = (td - timedelta(days=1)).strftime("%Y%m%d")
            start_d = (td - timedelta(days=10)).strftime("%Y%m%d")
            df_idx = pro.index_daily(start_date=start_d, end_date=end_d)
            if df_idx is not None and not df_idx.empty:
                df_idx = df_idx.sort_values("trade_date").groupby("ts_code").tail(1)
                data_source = "tushare_index_daily_prev_pre_market"
            else:
                df_idx = None
        elif is_intraday or (cur_minutes is None):
            # 盘中 (或 phase_time 缺失): 用 trade_date 当天
            df_idx = pro.index_daily(trade_date=td_str)
            if df_idx is None or df_idx.empty:
                # 降级: 取 trade_date 之前最近一个交易日
                start_d = (td - timedelta(days=10)).strftime("%Y%m%d")
                df_idx = pro.index_daily(start_date=start_d, end_date=td_str)
                if df_idx is not None and not df_idx.empty:
                    df_idx = df_idx.sort_values("trade_date").groupby("ts_code").tail(1)
                    data_source = "tushare_index_daily_prev"
            if is_intraday:
                data_source = "tushare_index_daily_intraday_use_open"
        else:
            # 盘后: 全量日线
            df_idx = pro.index_daily(trade_date=td_str)
            if df_idx is None or df_idx.empty:
                start_d = (td - timedelta(days=10)).strftime("%Y%m%d")
                df_idx = pro.index_daily(start_date=start_d, end_date=td_str)
                if df_idx is not None and not df_idx.empty:
                    df_idx = df_idx.sort_values("trade_date").groupby("ts_code").tail(1)
                    data_source = "tushare_index_daily_prev"

        # ── 优先用本地 index_1min parquet (反未来函数 + 零限速) ──
        from app.services.local_data_provider import local_data as _ld
        local_index_used = False
        local_index_fallback = False
        if is_intraday and hh is not None and mm is not None:
            for code in target_codes:
                win = _ld.get_index_minute_window(code, td, hh, mm)
                if win is not None:
                    local_index_used = True
                    name = basic_map.get(code, name_map.get(code, code))
                    open_change_pct = round((win["open"] - win["pre_close"]) / win["pre_close"] * 100, 2) if win["pre_close"] > 0 else 0.0
                    indices.append({
                        "symbol": code,
                        "name": name,
                        "current_price": win["current_price"],
                        "last_close": win["pre_close"],
                        "change": round(win["current_price"] - win["pre_close"], 2),
                        "change_pct": win["change_pct"],
                        "open": win["open"],
                        "high": win["high"],
                        "low": win["low"],
                        "volume": win["vol"],
                        "amount": win["amount"],
                        "trade_date": td_str,
                        "open_change_pct": open_change_pct,
                        "freshness": "intraday_minute",
                    })
            if local_index_used:
                data_source = "local_index_1min_intraday"
                return {
                    "indices": indices,
                    "trade_date": trade_date,
                    "data_source": data_source,
                    "freshness": "intraday_minute",
                    "caveat": None,
                    "count": len(indices),
                }
            # 本地无分钟数据, 降级到 Tushare 日线 (用 open 占位)
            local_index_fallback = True

        for code in target_codes:
            row = df_idx[df_idx["ts_code"] == code] if df_idx is not None else None
            if row is None or row.empty:
                continue
            r = row.iloc[0]
            name = basic_map.get(code, name_map.get(code, code))
            pre_close_val = float(r.get("pre_close", 0) or 0)
            open_val = float(r.get("open", 0) or 0)

            if is_intraday:
                # 盘中: 只能用 open 作为"当前价" (9:30 开盘价, 本地无指数分钟数据时)
                current_price = open_val
                change_val = 0.0
                change_pct_val = 0.0
                high_val = open_val
                low_val = open_val
                open_change_pct = round((open_val - pre_close_val) / pre_close_val * 100, 2) if pre_close_val > 0 else 0.0
                freshness = "intraday_open_only"
                caveat = f"⚠️ 本地无 {code} 指数分钟数据, 当前价为 9:30 开盘价占位 (开盘涨幅 {open_change_pct:+.2f}%)"
            elif is_pre_market:
                # 盘前: 用前一日收盘
                current_price = pre_close_val
                change_val = 0.0
                change_pct_val = 0.0
                high_val = pre_close_val
                low_val = pre_close_val
                open_change_pct = 0.0
                freshness = "pre_market"
                caveat = "⚠️ 盘前, 用前一日收盘价"
            else:
                # 盘后 / 缺 phase_time: 全量
                current_price = float(r.get("close", 0) or 0)
                change_val = float(r.get("change", 0) or 0)
                change_pct_val = float(r.get("pct_chg", 0) or 0)
                high_val = float(r.get("high", 0) or 0)
                low_val = float(r.get("low", 0) or 0)
                open_change_pct = round((float(r.get("open", 0) or 0) - pre_close_val) / pre_close_val * 100, 2) if pre_close_val > 0 else 0.0
                freshness = "post_market"
                caveat = None

            indices.append({
                "symbol": code,
                "name": name,
                "current_price": round(current_price, 2),
                "last_close": round(pre_close_val, 2),
                "change": round(change_val, 2),
                "change_pct": change_pct_val,
                "open": round(open_val, 2),
                "high": round(high_val, 2),
                "low": round(low_val, 2),
                "volume": float(r.get("vol", 0) or 0),
                "amount": float(r.get("amount", 0) or 0),
                "trade_date": str(r.get("trade_date", "")),
                # 真实已发生的开盘信息 (9:30 已知, 不算未来函数)
                "open_change_pct": open_change_pct,
                # 数据新鲜度标签
                "freshness": freshness,
            })

        return {
            "indices": indices,
            "trade_date": trade_date,
            "data_source": data_source,
            "freshness": freshness,  # 全局标签, 方便 Pi 一次判断
            "caveat": caveat,  # 给 Pi 的人类可读提示
            "count": len(indices),
        }
    except Exception as e:
        return {"indices": [], "error": f"Tushare index_daily 失败: {e}"}


@router.get("/{task_id}/sandbox/quote/{symbol}")
async def get_sandbox_quote(task_id: str, symbol: str,
                             trade_date: str = Query(None),
                             phase_time: str = Query(None, description="Pi 调用时刻 HH:MM（用于反查分钟/防未来函数）"),
                             hour: int = Query(None), minute: int = Query(None)):
    """返回回测时的瞬间行情快照（分钟级，按需懒加载单只标的）

    ⚠️ 反未来函数: 当 Pi 在盘中期(phase_time < 16:00)调用但分钟数据缺失时,
       不回退到当天日线 (那会泄漏收盘价), 而是返回 pre_close 占位
       (表示"今日数据未就绪, change_pct=0")
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta
    import random

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"error": "无法确定回测日期"}

    td = dt_date.fromisoformat(trade_date)

    # 优先用分钟数据（+0~3分钟随机延迟）
    if hour is not None and minute is not None:
        mq = local_data.get_minute_quote(symbol, td, hour, minute)
        if mq:
            pre_close = mq.get("pre_close", 0) or 0
            chg = round((mq["close"] - pre_close) / pre_close * 100, 2) if pre_close > 0 else 0
            # ── 累计日内高低点 (真正的日内分位, 而非单根分钟K线的分位) ──
            cumulative = local_data.get_cumulative_intraday_high_low(
                symbol, td, hour, minute
            )
            if cumulative:
                day_high = cumulative["day_high"]
                day_low = cumulative["day_low"]
                intraday_percentile = cumulative["intraday_percentile"]
            else:
                # 降级: 用分钟K线自身的高低点 (原行为, 可能不准确)
                day_high = mq["high"]
                day_low = mq["low"]
                intraday_percentile = None
            return {
                "symbol": symbol, "trade_date": trade_date,
                "source": "minute", "actual_time": mq["time"],
                "open": mq["open"],
                "high": day_high, "low": day_low,  # ← 累计日内高低点
                # 保留分钟K线原始 high/low 供调试
                "bar_high": mq["high"], "bar_low": mq["low"],
                "close": mq["close"], "pre_close": pre_close, "change_pct": chg,
                "volume": mq["volume"], "amount": mq["amount"],
                "intraday_percentile": intraday_percentile,
            }
        # 分钟数据缺失: 盘中期不允许回退日线 (会泄漏收盘价)
        if phase_time and _is_day_finalized(phase_time):
            pass  # 16:00+ 才允许走日线
        else:
            # 盘中期无分钟数据 → 返回 pre_close 占位 (Pi 视为"今日未开盘/数据未就绪")
            cur = td - timedelta(days=1)
            while cur.weekday() >= 5:
                cur -= timedelta(days=1)
            pre_close_only = local_data.get_daily_quote(symbol, cur)
            if pre_close_only:
                pre_close = float(pre_close_only.get("close", 0))
                return {
                    "symbol": symbol, "trade_date": trade_date,
                    "source": "pre_close_only",
                    "warning": f"⚠️ {trade_date} {phase_time} 无分钟数据, 返回昨日收盘价占位 (今日数据未就绪)",
                    "open": pre_close, "high": pre_close, "low": pre_close,
                    "close": pre_close, "pre_close": pre_close, "change_pct": 0.0,
                    "volume": 0, "amount": 0,
                }
            return {"error": f"{symbol} 无 {trade_date} 数据"}

    # 16:00+ / phase_time 缺失但 minute 缺失: 用日线 (此时 K 线已确认)
    q = local_data.get_daily_quote(symbol, td)
    if not q:
        return {"error": f"{symbol} 无 {trade_date} 数据"}
    # 日线路径: high/low 本身就是全天累计值, 直接计算分位
    day_high = q["high"]
    day_low = q["low"]
    close_p = q["close"]
    intraday_percentile = None
    if day_high > day_low and close_p > 0:
        intraday_percentile = round((close_p - day_low) / (day_high - day_low) * 100, 1)
    return {
        "symbol": symbol, "trade_date": trade_date,
        "source": "daily",
        "open": q["open"], "high": day_high, "low": day_low,
        "close": close_p, "pre_close": q["pre_close"],
        "change_pct": q["change_pct"], "volume": q["volume"], "amount": q["amount"],
        "intraday_percentile": intraday_percentile,
    }


@router.get("/{task_id}/sandbox/fibonacci/{symbol}")
async def get_sandbox_fibonacci(task_id: str, symbol: str,
                                 trade_date: str = Query(None),
                                 phase_time: str = Query(None)):
    """斐波那契回撤（基于本地日线），Pi 的 get_fibonacci_levels 回测路由到此
    16:00 前排除当天（未来函数防护）"""
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"error": "无法确定回测日期"}

    end_dt = dt_date.fromisoformat(trade_date)
    start_dt = end_dt - timedelta(days=120)

    include_today = _is_day_finalized(phase_time)
    bound = end_dt if include_today else (end_dt - timedelta(days=1))

    highs = []
    lows = []
    closes = []
    current = start_dt
    while current <= bound:
        q = local_data.get_daily_quote(symbol, current)
        if q:
            highs.append(q["high"])
            lows.append(q["low"])
            closes.append(q["close"])
        current += timedelta(days=1)

    if len(highs) < 10:
        return {"symbol": symbol, "error": "数据不足"}

    stage_high = max(highs)
    stage_low = min(lows)
    diff = stage_high - stage_low

    levels = [
        {"ratio": 0.0, "price": round(stage_low, 2), "label": "阶段底部"},
        {"ratio": 0.236, "price": round(stage_high - diff * 0.236, 2), "label": "弱支撑"},
        {"ratio": 0.382, "price": round(stage_high - diff * 0.382, 2), "label": "常规买点"},
        {"ratio": 0.5, "price": round(stage_high - diff * 0.5, 2), "label": "半分位"},
        {"ratio": 0.618, "price": round(stage_high - diff * 0.618, 2), "label": "强支撑/生死线"},
        {"ratio": 0.786, "price": round(stage_high - diff * 0.786, 2), "label": "深坑/放弃"},
        {"ratio": 1.0, "price": round(stage_high, 2), "label": "阶段顶部"},
    ]

    # current_price 用最后一条已确认 K 线收盘价
    current_price = closes[-1] if closes else stage_low

    # 判断当前区间
    zone = "观望区"
    suggestion = "等待回撤到位"
    for lv in levels:
        if current_price >= lv["price"] * 0.98:
            zone = lv["label"]
            if lv["ratio"] <= 0.382:
                suggestion = "接近底部区域，可考虑建仓"
            elif lv["ratio"] <= 0.618:
                suggestion = "处于合理回撤区域，观察确认"
            elif lv["ratio"] >= 0.786:
                suggestion = "接近顶部，注意止盈"
            break

    return {
        "symbol": symbol, "trade_date": trade_date,
        "high": stage_high, "low": stage_low, "diff": round(diff, 2),
        "current_price": current_price,
        "levels": levels,
        "position_zone": zone, "zone_suggestion": suggestion,
    }


@router.get("/{task_id}/sandbox/daily-channel/{symbol}")
async def get_sandbox_daily_channel(task_id: str, symbol: str,
                                    trade_date: str = Query(None),
                                    phase_time: str = Query(None)):
    """日内K值通道（基于本地日线），Pi 的 get_daily_channel 回测路由到此。
    - 16:00 之前: 用前日 K 线作为盘前通道 (避免未来函数)
    - 16:00 之后: 当天已收盘，使用当天 K 线
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    if not trade_date:
        db = SessionLocal()
        try:
            task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
            trade_date = task.current_day.isoformat() if task and task.current_day else None
        finally:
            db.close()
    if not trade_date:
        return {"error": "无法确定回测日期"}

    trade_dt = dt_date.fromisoformat(trade_date)
    include_today = _is_day_finalized(phase_time)

    if include_today:
        base_dt = trade_dt
    else:
        # 用前日 K 线
        base_dt = trade_dt - timedelta(days=1)
        while base_dt.weekday() >= 5:
            base_dt -= timedelta(days=1)

    q = local_data.get_daily_quote(symbol, base_dt)
    if not q:
        return {"symbol": symbol, "error": "无行情数据"}

    k = 0.98848
    avg_price = (q["high"] + q["low"] + q["close"]) / 3
    top_line = round(avg_price / k, 2)
    bottom_line = round(avg_price * k, 2)
    current = q["close"]
    channel_width = round((top_line - bottom_line) / bottom_line * 100, 2)

    return {
        "symbol": symbol, "trade_date": trade_date,
        "channel_base_date": base_dt.isoformat(),
        "include_today": include_today,
        "constant_k": k, "avg_price": round(avg_price, 2),
        "current_price": current,
        "top_line": top_line, "bottom_line": bottom_line,
        "channel_width_pct": channel_width,
        "position": "上方" if current > top_line else ("下方" if current < bottom_line else "通道内"),
    }


# ════════════════════════════════════════════════════════════════════════
# 沙盒版 calc_position：所有 Layer1 数据走本地（避免实时接口 + 未来函数）
# ════════════════════════════════════════════════════════════════════════

@router.post("/{task_id}/sandbox/calc-position")
async def calc_position_sandbox(task_id: str, req: dict):
    """回测模式下的仓位计算
    数据源:
      - 当前价/涨跌幅: sandbox/quote (分钟级 or 日线)
      - 振幅: sandbox/kline 前 5 日 (排除当天未来函数)
      - 大盘指数: sandbox/indices (本地 index_basic + Tushare index_daily, 自动降级到前一日)
      - 账户/持仓: 与实时一致, 从沙盒引擎 (BacktestPaperEngine) 取
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date
    from app.api.indicator import (
        _get_single_stock_cap, _get_role_cap, _get_total_cap,
        _get_tier_condition, _get_tier_profit_threshold,
        _get_amplitude_tier, _get_dynamic_stop_pct, _get_iron_rule2,
        _calculate_amplitude_from_kline,
    )

    # ── 解析参数 ──
    symbol = (req.get("symbol") or "").strip().upper()
    if not symbol:
        return {"error": "缺少 symbol 参数"}
    signal_strength = req.get("signal_strength", "medium")
    chain_role = req.get("chain_role", "mid")
    tier = req.get("tier", "probe")
    stance = req.get("stance", "yellow")
    phase_time = req.get("phase_time")  # 来自 [BKT:HH:MM] 前缀

    # ── 1. 获取任务日期 ──
    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            return {"error": f"任务 {task_id} 不存在"}
        trade_date = task.current_day
        if not trade_date:
            return {"error": "任务无当前日期"}
        trade_date_str = trade_date.isoformat()
    finally:
        db.close()

    warnings: list = []

    # ── 2. 沙盒账户信息 ──
    from app.core.trading.backtest_paper import BacktestPaperEngine
    try:
        engine = BacktestPaperEngine(task_id, initial_capital=task.initial_capital or 1000000)
        # 走 get_account() 拿标准 dict（BacktestPaperEngine 没有 available_cash/frozen_cash 直接属性）
        acc = engine.get_account(day_df=None)  # 用 avg_price 估值（不引入未来行情）
        available_cash = float(acc.get("available_cash", 0))
        frozen_cash = float(acc.get("frozen_cash", 0))
        # 取持仓列表（BacktestPaperEngine.get_positions() 返回标准 dict 列表）
        position_list = engine.get_positions()
        position_value = 0.0
        for pos in position_list:
            sym = pos.get("symbol") or ""
            try:
                q = local_data.get_daily_quote(sym, trade_date)
                price = float(q["close"]) if q else float(pos.get("avg_cost", 0))
            except Exception:
                price = float(pos.get("avg_cost", 0))
            position_value += float(pos.get("volume", 0)) * price
        total_asset = available_cash + frozen_cash + position_value
        position_ratio = round(position_value / total_asset * 100, 2) if total_asset > 0 else 0
    except Exception as e:
        logger.warning(f"[sandbox/calc-position] 沙盒账户读取失败: {e}")
        total_asset = float(task.initial_capital or 1000000)
        available_cash = total_asset
        position_value = 0.0
        position_ratio = 0.0
        position_list = []
        warnings.append(f"⚠️ 沙盒账户读取失败: {e}")

    # ── 3. 当前价 + 涨跌幅 (sandbox/quote) ──
    current_price = 0.0
    change_pct = 0.0
    stock_name = ""
    hh, mm = (None, None)
    if phase_time:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
        except Exception:
            pass
    try:
        qres = await get_sandbox_quote(
            task_id=task_id, symbol=symbol, trade_date=trade_date_str,
            hour=hh, minute=mm,
        )
        if qres.get("error"):
            warnings.append(f"⚠️ {qres['error']}")
        else:
            current_price = float(qres.get("close", 0))
            change_pct = float(qres.get("change_pct", 0))
            pre_close = float(qres.get("pre_close", 0))
            if pre_close <= 0 and "pre_close" not in qres:
                # 分钟快照里 pre_close 已补过; 日线分支也有
                pass
    except Exception as e:
        warnings.append(f"⚠️ 当前价拉取失败: {e}")

    if current_price <= 0:
        return {"symbol": symbol, "error": f"无法获取 {symbol} 在 {trade_date_str} {phase_time or ''} 的价格"}

    # ── 4. 近 5 日振幅 (sandbox/kline, 含 phase_time 未来函数防护) ──
    amplitude = 0.0
    try:
        kres = await get_sandbox_kline(
            task_id=task_id, symbol=symbol, limit=5,
            trade_date=trade_date_str, phase_time=phase_time,
        )
        klines = kres.get("kline", [])
        if klines:
            kobj_list = [type("K", (), {
                "close": float(k.get("close", 0) or 0),
                "high": float(k.get("high", 0) or 0),
                "low": float(k.get("low", 0) or 0),
            }) for k in klines]
            amplitude = _calculate_amplitude_from_kline(kobj_list)
    except Exception as e:
        warnings.append(f"⚠️ 振幅计算失败: {e}")

    # ── 5. 大盘指数涨跌幅 (sandbox/indices, 已含未来函数防护: 非交易日自动降级) ──
    index_pct = 0.0
    try:
        ires = await get_sandbox_indices(task_id=task_id, trade_date=trade_date_str)
        for idx in ires.get("indices", []):
            if idx.get("symbol") == "000001.SH":  # 上证指数
                index_pct = float(idx.get("change_pct", 0))
                break
        if index_pct == 0.0 and ires.get("indices"):
            index_pct = float(ires["indices"][0].get("change_pct", 0))
    except Exception as e:
        warnings.append(f"⚠️ 大盘指数拉取失败: {e}")

    # ── 6. 加载约束条件 + 计算仓位 (复用 indicator.py 纯函数) ──
    single_cap_pct = _get_single_stock_cap(signal_strength)
    role_cap_pct = _get_role_cap(chain_role)
    total_cap_pct = _get_total_cap(stance)
    tier_condition = _get_tier_condition(tier)
    amplitude_tier = _get_amplitude_tier(amplitude)
    dynamic_stop_pct = _get_dynamic_stop_pct(index_pct, amplitude)

    def _round_lot(shares: float) -> int:
        return max(0, int(shares // 100) * 100)

    effective_single_cap = min(single_cap_pct, role_cap_pct) / 100.0 * total_asset
    total_remaining = total_cap_pct / 100.0 * total_asset - position_value
    cash_reserve_line = total_asset * 0.25
    cash_available_for_buy = available_cash - cash_reserve_line
    max_usable = min(effective_single_cap, max(total_remaining, 0), max(cash_available_for_buy, 0))

    max_shares = _round_lot(max_usable / current_price)
    max_amount = round(max_shares * current_price, 2)
    rec_amount_raw = min(role_cap_pct / 100.0 * total_asset, max_usable)
    rec_shares = _round_lot(rec_amount_raw / current_price)
    rec_amount = round(rec_shares * current_price, 2)
    rec_pct = round(rec_amount / total_asset * 100, 2) if total_asset > 0 else 0
    probe_amount_raw = min(0.10 * total_asset, max_usable)
    probe_shares = _round_lot(probe_amount_raw / current_price)
    probe_amount = round(probe_shares * current_price, 2)
    probe_pct = round(probe_amount / total_asset * 100, 2) if total_asset > 0 else 0

    # ── 7. 止损 ──
    hard_stop_price = round(current_price * (1 - dynamic_stop_pct / 100.0), 3)
    max_loss_per_share = round(current_price - hard_stop_price, 3)
    total_max_loss = round(rec_shares * max_loss_per_share, 2)
    iron_rule = _get_iron_rule2(amplitude_tier)

    # ── 8. 逐条验证 ──
    single_cap_actual_pct = round(rec_amount / total_asset * 100, 2) if total_asset > 0 else 0
    single_cap_ok = rec_amount <= effective_single_cap
    single_cap_detail = f"建议{rec_amount}({single_cap_actual_pct}%) ≤ 上限{round(effective_single_cap,2)}({min(single_cap_pct, role_cap_pct)}%)"
    new_total_position = position_value + rec_amount
    new_total_ratio = round(new_total_position / total_asset * 100, 2) if total_asset > 0 else 0
    total_position_ok = new_total_position <= total_cap_pct / 100.0 * total_asset
    total_position_detail = f"建仓后{new_total_ratio}% ≤ 上限{total_cap_pct}%"
    new_cash = available_cash - rec_amount
    new_cash_ratio = round(new_cash / total_asset * 100, 2) if total_asset > 0 else 0
    cash_reserve_ok = new_cash >= cash_reserve_line
    cash_reserve_detail = f"建仓后现金{new_cash_ratio}% ≥ 底线25%"
    loss_ratio = round(total_max_loss / total_asset * 100, 2) if total_asset > 0 else 0
    max_loss_ok = total_max_loss <= total_asset * 0.02
    max_loss_detail = f"单笔亏损{total_max_loss}({loss_ratio}%) ≤ 上限{round(total_asset*0.02,2)}(2%)"

    pre_condition_ok = None
    pre_condition_detail = None
    threshold = _get_tier_profit_threshold(tier)
    if threshold > 0:
        existing_profit = None
        try:
            for p in position_list:
                pos = p if isinstance(p, dict) else p.__dict__
                p_sym = (pos.get("symbol") or pos.get("ts_code") or "").upper()
                p_sym_norm = p_sym.replace("SH", "").replace("SZ", "").replace("BJ", "")
                sym_norm = symbol.replace("SH", "").replace("SZ", "").replace("BJ", "")
                if p_sym_norm == sym_norm:
                    avg_price = float(pos.get("avg_price", 0))
                    if avg_price > 0:
                        existing_profit = round((current_price - avg_price) / avg_price * 100, 2)
                        break
        except Exception:
            pass
        if existing_profit is None:
            pre_condition_ok = False
            pre_condition_detail = f"无该股持仓，{tier}层级需要已有持仓且浮盈≥{threshold}%"
            warnings.append(f"⚠️ 前仓条件不满足: {tier}层级需要已有该股持仓且浮盈≥{threshold}%")
        elif existing_profit < threshold:
            pre_condition_ok = False
            pre_condition_detail = f"当前浮盈{existing_profit}% < 所需{threshold}%"
            warnings.append(f"⚠️ 前仓条件不满足: 当前浮盈{existing_profit}% < 所需{threshold}%")
        else:
            pre_condition_ok = True
            pre_condition_detail = f"当前浮盈{existing_profit}% ≥ 所需{threshold}%"

    all_pass = single_cap_ok and total_position_ok and cash_reserve_ok and max_loss_ok
    if pre_condition_ok is not None and not pre_condition_ok:
        all_pass = False

    from app.models.indicator import (
        CalcPositionQuantity, CalcPositionStopLoss,
        CalcPositionValidation, CalcPositionResponse,
    )

    return CalcPositionResponse(
        symbol=symbol, name=stock_name,
        total_asset=round(total_asset, 2),
        available_cash=round(available_cash, 2),
        position_value=round(position_value, 2),
        position_ratio=position_ratio,
        signal_strength=signal_strength,
        single_stock_cap_pct=single_cap_pct,
        chain_role=chain_role,
        role_cap_pct=role_cap_pct,
        tier=tier, tier_condition=tier_condition,
        stance=stance, total_cap_pct=total_cap_pct,
        amplitude=round(amplitude, 2), amplitude_tier=amplitude_tier,
        index_pct=round(index_pct, 2),
        current_price=current_price,
        quantity=CalcPositionQuantity(
            max_shares=max_shares, max_amount=max_amount,
            rec_shares=rec_shares, rec_amount=rec_amount, rec_pct=rec_pct,
            probe_shares=probe_shares, probe_amount=probe_amount, probe_pct=probe_pct,
        ),
        stop_loss=CalcPositionStopLoss(
            volatility_tier=amplitude_tier,
            dynamic_stop_pct=dynamic_stop_pct,
            hard_stop_price=hard_stop_price,
            max_loss_per_share=max_loss_per_share,
            total_max_loss=total_max_loss,
            iron_rule2_t1_pct=iron_rule["t1_pct"],
            iron_rule2_t2_pct=iron_rule["t2_pct"],
            iron_rule2_t2_plus_pct=iron_rule["t2_plus_pct"],
            iron_rule2_t3_pct=iron_rule["t3_pct"],
            iron_rule2_t3_plus_pct=iron_rule["t3_plus_pct"],
        ),
        validation=CalcPositionValidation(
            single_cap_ok=single_cap_ok, single_cap_detail=single_cap_detail,
            total_position_ok=total_position_ok, total_position_detail=total_position_detail,
            cash_reserve_ok=cash_reserve_ok, cash_reserve_detail=cash_reserve_detail,
            max_loss_ok=max_loss_ok, max_loss_detail=max_loss_detail,
            pre_condition_ok=pre_condition_ok, pre_condition_detail=pre_condition_detail,
        ),
        warnings=warnings,
        data_source="backtest_local",
        trade_date=trade_date_str,
        phase_time=phase_time,
    ).dict()


# ═══════════════════════════════════════════════════════════════
# 条件诊断端点：加仓条件 / 建仓条件
# ═══════════════════════════════════════════════════════════════

@router.get("/{task_id}/sandbox/position-add-conditions")
async def position_add_conditions(
    task_id: str,
    symbol: str = Query(None, description="指定标的，不传则检查全部持仓"),
    phase_time: str = Query(None, description="阶段时间 HH:MM，用于获取盘中价格"),
):
    """查询持仓距离加仓还差哪些条件。

    逐只检查当前持仓：
      - 当前层级 + 浮盈 → 是否触发层级升级评估
      - 6 道门控逐一判定（Pi 立场 / 回撤 / 保护线 / 日加仓上限 / 趋势强度）
      - 返回每道门的通过/未通过状态及未通过原因
    """
    from app.services.local_data_provider import local_data
    from app.core.trading.backtest_paper import BacktestPaperEngine
    from datetime import date as dt_date, timedelta

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, f"任务 {task_id} 不存在")
        trade_date = task.current_day
        if not trade_date:
            return {"error": "任务尚未开始或无当前日期", "task_id": task_id}
        trade_date_str = trade_date.isoformat()
        initial_capital = float(task.initial_capital or 1000000)
    finally:
        db.close()

    hh = mm = None
    if phase_time:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
        except Exception:
            pass

    # ── 获取沙盒账户 + 持仓 ──
    try:
        engine = BacktestPaperEngine(task_id, initial_capital=initial_capital)
        acc = engine.get_account(day_df=None)
        positions = engine.get_positions()
        total_asset = float(acc.get("total_asset", initial_capital))
        available_cash = float(acc.get("available_cash", 0))
    except Exception as e:
        return {"error": f"沙盒账户读取失败: {e}", "task_id": task_id}

    if not positions:
        return {
            "task_id": task_id,
            "trade_date": trade_date_str,
            "total_asset": total_asset,
            "positions": [],
            "summary": "当前无持仓",
        }

    # ── Pi 立场（从最近日志读取）──
    pi_stance = "yellow"
    try:
        db2 = SessionLocal()
        recent_log = (
            db2.query(BacktestDailyLog)
            .filter(
                BacktestDailyLog.task_id == task_id,
                BacktestDailyLog.log_type == "pi_decision",
            )
            .order_by(BacktestDailyLog.created_at.desc())
            .first()
        )
        if recent_log and recent_log.content:
            import re as _re
            m = _re.search(r'立场[：:]\s*(green|yellow|red)', str(recent_log.content))
            if m:
                pi_stance = m.group(1)
        db2.close()
    except Exception:
        pass

    # ── 总回撤 ──
    total_pnl = float(acc.get("float_pnl", 0)) + float(acc.get("realized_pnl", 0))
    total_drawdown = -min(0, total_pnl) / initial_capital if initial_capital > 0 else 0

    # ── 逐只评估 ──
    results = []
    for pos in positions:
        sym = pos.get("symbol", "")
        if not sym or (symbol and sym != symbol):
            continue

        volume = int(pos.get("volume", 0))
        avg_cost = float(pos.get("avg_cost", pos.get("avg_price", 0)))
        if volume <= 0 or avg_cost <= 0:
            continue

        # 当前价（盘中快照或日线收盘价）
        cur_price = 0.0
        if hh is not None and mm is not None:
            q = local_data.get_minute_quote(sym, trade_date, hh, mm)
            if q:
                cur_price = float(q.get("close", 0))
        if cur_price <= 0:
            bar = local_data.get_daily_quote(sym, trade_date)
            if bar:
                cur_price = float(bar.get("close", 0))
        if cur_price <= 0:
            continue

        float_pnl = (cur_price / avg_cost) - 1
        float_pnl_pct = round(float_pnl * 100, 2)

        # T+1 锁定
        t1 = engine.get_t1_status(sym) if hasattr(engine, "get_t1_status") else {}
        t1_locked = t1.get("locked", False)

        # ── 层级评估 ──
        current_tier = "probe"  # backtest 默认从 probe 开始
        tier_info = _eval_position_tier_diag(float_pnl, current_tier)

        # ── 门控检查 ──
        gates = _eval_position_gates_diag(
            symbol=sym,
            float_pnl=float_pnl,
            target_tier=tier_info.get("target_tier", ""),
            trade_date=trade_date,
            total_asset=total_asset,
            total_drawdown=total_drawdown,
            pi_stance=pi_stance,
            t1_locked=t1_locked,
            current_price=cur_price,
            avg_cost=avg_cost,
        )

        # 汇总缺失条件
        missing_conditions = []
        if t1_locked:
            missing_conditions.append("T+1锁定中，今日买入无法加仓")
        elif tier_info.get("action") == "MAX_TIER":
            missing_conditions.append("已达最高层级(冲刺仓)，无法继续加仓")
        elif tier_info.get("action") == "HOLD":
            need = tier_info.get("need_pnl", 0)
            missing_conditions.append(
                f"浮盈不足: 当前{float_pnl_pct}%, 需要≥{round(need*100,1)}%才能触发{tier_info.get('next_tier','')}升级"
            )
        else:
            for g in gates:
                if not g["passed"]:
                    missing_conditions.append(g["detail"])

        current_mv = cur_price * volume
        current_pct = round(current_mv / total_asset * 100, 2) if total_asset > 0 else 0

        results.append({
            "symbol": sym,
            "current_price": round(cur_price, 2),
            "avg_cost": round(avg_cost, 2),
            "volume": volume,
            "market_value": round(current_mv, 2),
            "position_pct": current_pct,
            "float_pnl_pct": float_pnl_pct,
            "current_tier": current_tier,
            "t1_locked": t1_locked,
            "tier_evaluation": tier_info,
            "gates": gates,
            "missing_conditions": missing_conditions,
            "can_add": len(missing_conditions) == 0,
        })

    return {
        "task_id": task_id,
        "trade_date": trade_date_str,
        "phase_time": phase_time,
        "total_asset": round(total_asset, 2),
        "available_cash": round(available_cash, 2),
        "total_drawdown_pct": round(total_drawdown * 100, 2),
        "pi_stance": pi_stance,
        "positions": results,
        "summary": (
            f"共{len(results)}只持仓, "
            f"{sum(1 for r in results if r['can_add'])}只满足加仓条件, "
            f"{sum(1 for r in results if not r['can_add'])}只条件不足"
        ),
    }


@router.get("/{task_id}/sandbox/candidate-entry-conditions")
async def candidate_entry_conditions(
    task_id: str,
    symbol: str = Query(None, description="指定标的，不传则检查全部候选池标的"),
    phase_time: str = Query(None, description="阶段时间 HH:MM"),
):
    """查询候选池股票距离建仓还差哪些条件。

    逐只检查候选池中 waiting 状态的标的：
      - 入场过滤三层（技术面 / 主力资金 / 超买）
      - Pi 立场检查
      - 午后额外检查（涨幅/分位）
      - 基本仓位验证
      - 返回每层的通过/未通过状态及具体原因
    """
    from app.services.local_data_provider import local_data
    from datetime import date as dt_date, timedelta

    db = SessionLocal()
    try:
        task = db.query(BacktestTask).filter(BacktestTask.id == task_id).first()
        if not task:
            raise HTTPException(404, f"任务 {task_id} 不存在")
        trade_date = task.current_day
        if not trade_date:
            return {"error": "任务尚未开始或无当前日期", "task_id": task_id}
        trade_date_str = trade_date.isoformat()
        initial_capital = float(task.initial_capital or 1000000)
    finally:
        db.close()

    hh = mm = None
    if phase_time:
        try:
            hh, mm = map(int, phase_time.split(":")[:2])
        except Exception:
            pass

    # ── 获取沙盒账户 ──
    try:
        from app.core.trading.backtest_paper import BacktestPaperEngine
        engine = BacktestPaperEngine(task_id, initial_capital=initial_capital)
        acc = engine.get_account(day_df=None)
        total_asset = float(acc.get("total_asset", initial_capital))
        available_cash = float(acc.get("available_cash", 0))
        positions = engine.get_positions()
        position_value = 0.0
        for p in positions:
            sym = p.get("symbol", "")
            vol = float(p.get("volume", 0))
            cost = float(p.get("avg_cost", p.get("avg_price", 0)))
            try:
                q = local_data.get_daily_quote(sym, trade_date)
                px = float(q["close"]) if q else cost
            except Exception:
                px = cost
            position_value += vol * px
    except Exception as e:
        return {"error": f"沙盒账户读取失败: {e}", "task_id": task_id}

    # ── Pi 立场 ──
    pi_stance = "yellow"
    try:
        db2 = SessionLocal()
        recent_log = (
            db2.query(BacktestDailyLog)
            .filter(
                BacktestDailyLog.task_id == task_id,
                BacktestDailyLog.log_type == "pi_decision",
            )
            .order_by(BacktestDailyLog.created_at.desc())
            .first()
        )
        if recent_log and recent_log.content:
            import re as _re
            m = _re.search(r'立场[：:]\s*(green|yellow|red)', str(recent_log.content))
            if m:
                pi_stance = m.group(1)
        db2.close()
    except Exception:
        pass

    # ── 获取候选池 ──
    from app.services.candidate_pool import get_candidate_pool
    pool = get_candidate_pool()
    waiting = pool.get_waiting()
    if symbol:
        waiting = [c for c in waiting if c.get("symbol") == symbol]
    if not waiting:
        return {
            "task_id": task_id,
            "trade_date": trade_date_str,
            "candidates": [],
            "summary": "候选池无 waiting 状态标的",
        }

    # ── 午后判断 ──
    is_afternoon = hh is not None and hh >= 13

    # ── 逐只评估 ──
    results = []
    for candidate in waiting:
        sym = candidate.get("symbol", "")
        name = candidate.get("name", "")

        # 获取当前价
        cur_price = 0.0
        prev_close = 0.0
        if hh is not None and mm is not None:
            q = local_data.get_minute_quote(sym, trade_date, hh, mm)
            if q:
                cur_price = float(q.get("close", 0))
        if cur_price <= 0:
            bar = local_data.get_daily_quote(sym, trade_date)
            if bar:
                cur_price = float(bar.get("close", 0))
                prev_close = float(bar.get("pre_close", 0))
        if cur_price <= 0:
            results.append({
                "symbol": sym,
                "name": name,
                "error": "无法获取当前价格",
                "missing_conditions": ["无法获取当前价格"],
                "can_entry": False,
            })
            continue

        if prev_close <= 0:
            # 从日线获取前收盘
            bar = local_data.get_daily_quote(sym, trade_date)
            if bar:
                prev_close = float(bar.get("pre_close", 0))
        change_pct = round((cur_price / prev_close - 1) * 100, 2) if prev_close > 0 else 0

        # ── 搜集技术指标数据 ──
        closes = _get_recent_closes(sym, trade_date, cur_price, local_data)
        ma5 = sum(closes[-5:]) / 5 if len(closes) >= 5 else 0
        ma20 = sum(closes[-20:]) / 20 if len(closes) >= 20 else 0
        rsi6 = _calc_rsi(closes, 6) if len(closes) >= 7 else 50

        # ── Layer 1: 技术面 ──
        above_ma5 = cur_price > ma5 if ma5 > 0 else False
        above_ma20 = cur_price > ma20 if ma20 > 0 else False
        ma5_gt_ma20 = ma5 > ma20 if ma5 > 0 and ma20 > 0 else False
        l1_pass = above_ma5 and above_ma20
        if not ma5_gt_ma20:
            l1_pass = False  # 死叉
        l1_failed = []
        if not above_ma5:
            l1_failed.append(f"价格{cur_price:.2f} ≤ MA5({ma5:.2f})" if ma5 > 0 else "MA5数据不足")
        if not above_ma20:
            l1_failed.append(f"价格{cur_price:.2f} ≤ MA20({ma20:.2f})" if ma20 > 0 else "MA20数据不足")
        if not ma5_gt_ma20:
            l1_failed.append(f"MA5({ma5:.2f}) ≤ MA20({ma20:.2f}) 死叉" if ma5 > 0 and ma20 > 0 else "MA排列数据不足")
        if l1_pass and not l1_failed:
            l1_failed = []

        # ── Layer 2: 主力资金 ──
        main_net_5d, main_net_5d_detail = _get_main_net_5d_diag(sym, trade_date, hh, mm, local_data)
        l2_pass = main_net_5d > 0
        l2_failed = []
        if not l2_pass:
            l2_failed.append(f"5日主力净流入{main_net_5d/1e8:.2f}亿 ≤ 0")

        # ── Layer 3: 超买 ──
        l3_pass = rsi6 < 70
        l3_failed = []
        if not l3_pass:
            l3_failed.append(f"RSI6={rsi6:.1f} ≥ 70 超买")

        # ── 综合过滤结果 ──
        all_filter_pass = l1_pass and l2_pass and l3_pass

        # ── 涨幅买入确认 ──
        buy_confirm_pass = True
        buy_confirm_failed = []
        if change_pct > 8:
            buy_confirm_pass = False
            buy_confirm_failed.append(f"涨幅{change_pct}% > 8%，放弃追涨")
        elif change_pct > 5:
            buy_confirm_pass = False
            buy_confirm_failed.append(f"涨幅{change_pct}% > 5% 且 ≤ 8%，需量比>1.5确认")
        # 回测简化: change_pct <= 5 直接通过

        # ── PT立场检查 ──
        stance_pass = pi_stance != "red"
        stance_detail = f"Pi立场={pi_stance}" + (" (red禁止建仓)" if not stance_pass else " (允许)")

        # ── 午后检查 ──
        afternoon_pass = True
        afternoon_failed = []
        if is_afternoon:
            if change_pct > 3:
                afternoon_pass = False
                afternoon_failed.append(f"午后涨幅{change_pct}% > 3%")
            # 日内分位检查
            cum = None
            if hh is not None and mm is not None:
                try:
                    cum = local_data.get_cumulative_intraday_high_low(sym, trade_date, hh, mm)
                except Exception:
                    pass
            intraday_pct = cum.get("intraday_percentile") if cum else None
            if intraday_pct is not None and intraday_pct > 60:
                afternoon_pass = False
                afternoon_failed.append(f"午后分位{intraday_pct:.0f}% > 60%")

        # ── 汇总 ──
        missing_conditions = []
        if not l1_pass:
            missing_conditions.append(f"[技术面] {', '.join(l1_failed)}")
        if not l2_pass:
            missing_conditions.append(f"[主力资金] {', '.join(l2_failed)}")
        if not l3_pass:
            missing_conditions.append(f"[超买] {', '.join(l3_failed)}")
        if not stance_pass:
            missing_conditions.append(f"[立场] {stance_detail}")
        if not buy_confirm_pass:
            missing_conditions.append(f"[买入确认] {', '.join(buy_confirm_failed)}")
        if not afternoon_pass:
            missing_conditions.append(f"[午后限制] {', '.join(afternoon_failed)}")

        can_entry = all_filter_pass and stance_pass and buy_confirm_pass and afternoon_pass

        # 候选池周期判断（长期/短期）
        added_date = candidate.get("added_trade_day", "")
        pool_type = "短期"  # default
        try:
            if added_date:
                added_dt = dt_date.fromisoformat(added_date)
                days_in_pool = (trade_date - added_dt).days
                pool_type = "短期" if days_in_pool <= 3 else "长期"
        except Exception:
            pass

        results.append({
            "symbol": sym,
            "name": name,
            "pool_type": pool_type,
            "added_date": added_date,
            "current_price": round(cur_price, 2),
            "change_pct": change_pct,
            "last_reject_reason": candidate.get("last_reject_reason", candidate.get("reject_reasons", [])),
            "checks_count": candidate.get("checks_count", 0),
            "filters": {
                "layer1_tech": {
                    "passed": l1_pass,
                    "above_ma5": above_ma5, "above_ma20": above_ma20,
                    "ma5_gt_ma20": ma5_gt_ma20,
                    "ma5": round(ma5, 2), "ma20": round(ma20, 2),
                    "failed": l1_failed,
                },
                "layer2_capital": {
                    "passed": l2_pass,
                    "main_net_5d_e8": round(main_net_5d / 1e8, 2),
                    "detail": main_net_5d_detail,
                    "failed": l2_failed,
                },
                "layer3_overbought": {
                    "passed": l3_pass,
                    "rsi6": round(rsi6, 1),
                    "failed": l3_failed,
                },
            },
            "stance_check": {"passed": stance_pass, "detail": stance_detail},
            "buy_confirmation": {"passed": buy_confirm_pass, "failed": buy_confirm_failed},
            "afternoon_check": {"passed": afternoon_pass, "failed": afternoon_failed, "is_afternoon": is_afternoon},
            "missing_conditions": missing_conditions,
            "can_entry": can_entry,
        })

    long_term = [r for r in results if r["pool_type"] == "长期"]
    short_term = [r for r in results if r["pool_type"] == "短期"]

    return {
        "task_id": task_id,
        "trade_date": trade_date_str,
        "phase_time": phase_time,
        "total_asset": round(total_asset, 2),
        "available_cash": round(available_cash, 2),
        "position_value": round(position_value, 2),
        "pi_stance": pi_stance,
        "is_afternoon": is_afternoon,
        "candidates": results,
        "summary": (
            f"长期候选池 {len(long_term)} 只 / 短期候选池 {len(short_term)} 只, "
            f"共 {sum(1 for r in results if r['can_entry'])} 只满足建仓条件, "
            f"{sum(1 for r in results if not r['can_entry'])} 只条件不足"
        ),
        "long_term": long_term,
        "short_term": short_term,
    }


# ═══════════════════════════════════════════════════════════════
# 诊断辅助函数
# ═══════════════════════════════════════════════════════════════

def _eval_position_tier_diag(float_pnl: float, current_tier: str) -> dict:
    """诊断用层级评估，返回详细信息"""
    if current_tier == "sprint":
        return {
            "action": "MAX_TIER",
            "current_tier": "sprint",
            "target_tier": None,
            "next_tier": None,
            "signal": "已达冲刺仓（最高层级），无法继续加仓",
            "need_pnl": None,
        }
    if float_pnl >= 0.03:
        return {
            "action": "UPGRADE_TO_SPRINT",
            "current_tier": current_tier,
            "target_tier": "sprint",
            "next_tier": "sprint",
            "max_position_pct": 0.25,
            "signal": f"浮盈{float_pnl:.1%} ≥ 3%，触发冲刺仓评估",
            "need_pnl": None,
        }
    if current_tier in ("probe", "unknown", "") and float_pnl >= 0.01:
        return {
            "action": "UPGRADE_TO_CONFIRM",
            "current_tier": "probe",
            "target_tier": "confirm",
            "next_tier": "confirm",
            "max_position_pct": 0.18,
            "signal": f"浮盈{float_pnl:.1%} ≥ 1%，触发确认仓评估",
            "need_pnl": None,
        }
    # HOLD: 浮盈不足
    if current_tier in ("probe", "unknown", ""):
        next_tier = "confirm"
        need = 0.01
    else:
        next_tier = "sprint"
        need = 0.03
    return {
        "action": "HOLD",
        "current_tier": current_tier,
        "target_tier": None,
        "next_tier": next_tier,
        "signal": f"浮盈{float_pnl:.1%} 不满足升级条件",
        "need_pnl": need,
    }


def _eval_position_gates_diag(
    symbol: str, float_pnl: float, target_tier: str,
    trade_date, total_asset: float, total_drawdown: float,
    pi_stance: str, t1_locked: bool,
    current_price: float, avg_cost: float,
) -> list:
    """诊断用门控检查，返回每道门的通过状态和详情"""
    from datetime import date as dt_date, timedelta
    from app.services.local_data_provider import local_data

    gates = []

    # 门控 1: Pi 立场
    if pi_stance == "red":
        total_pos_pct = 0  # simplified
        if total_pos_pct + 0.18 > 0.20:  # rough check
            gates.append({
                "name": "Pi立场",
                "passed": False,
                "detail": f"RED立场下总仓超20%限制，禁止加仓",
            })
        else:
            gates.append({
                "name": "Pi立场",
                "passed": True,
                "detail": "RED例外：验证盈利头寸限额内",
            })
    else:
        gates.append({
            "name": "Pi立场",
            "passed": True,
            "detail": f"{pi_stance.upper()}立场通过",
        })

    # 门控 2: 总回撤
    if total_drawdown >= 0.05:
        gates.append({
            "name": "总回撤",
            "passed": False,
            "detail": f"总回撤{total_drawdown:.1%} ≥ 5%，硬禁止加仓",
        })
    else:
        gates.append({
            "name": "总回撤",
            "passed": True,
            "detail": f"回撤{total_drawdown:.1%} < 5%，通过",
        })

    # 门控 3: 保护线
    if target_tier in ("confirm", "sprint"):
        amp = _calc_avg_amplitude_diag(symbol, trade_date, local_data)
        if target_tier == "confirm":
            if float_pnl < 0:
                gates.append({
                    "name": "保护线(T1)",
                    "passed": False,
                    "detail": f"T1保本线：浮盈{float_pnl:.1%} < 0，已跌破成本价",
                })
            elif float_pnl < 0.005:
                gates.append({
                    "name": "保护线(T1)",
                    "passed": False,
                    "detail": f"T1保本线：浮盈{float_pnl:.1%}距离保本线不足0.5%",
                })
            else:
                gates.append({
                    "name": "保护线(T1)",
                    "passed": True,
                    "detail": f"T1保本线：浮盈{float_pnl:.1%} ≥ 0.5%，安全",
                })
        else:  # sprint
            if amp <= 0.03:
                x, tier_label = 0.01, "低波"
            elif amp <= 0.06:
                x, tier_label = 0.02, "中波"
            else:
                x, tier_label = 0.03, "高波"
            if float_pnl < x:
                gates.append({
                    "name": "保护线(T2)",
                    "passed": False,
                    "detail": f"T2保护线({tier_label})：浮盈{float_pnl:.1%} < 成本+{x:.0%}",
                })
            elif float_pnl - x < 0.005:
                gates.append({
                    "name": "保护线(T2)",
                    "passed": False,
                    "detail": f"T2保护线({tier_label})：浮盈{float_pnl:.1%}距离保护线不足0.5%",
                })
            else:
                gates.append({
                    "name": "保护线(T2)",
                    "passed": True,
                    "detail": f"T2保护线({tier_label})：浮盈{float_pnl:.1%} ≥ 成本+{x:.0%}，安全",
                })
    else:
        gates.append({
            "name": "保护线",
            "passed": True,
            "detail": "未触发层级升级，保护线不适用",
        })

    # 门控 4: T+1 锁定
    if t1_locked:
        gates.append({
            "name": "T+1锁定",
            "passed": False,
            "detail": "今日买入，T+1锁定中，无法加仓",
        })
    else:
        gates.append({
            "name": "T+1锁定",
            "passed": True,
            "detail": "非今日买入，可以加仓",
        })

    # 门控 5: 趋势强度（简化：MA5斜率 + 量比 + MA5>MA20 + 资金流向）
    trend = _check_trend_strength_diag(symbol, trade_date, local_data)
    gates.append({
        "name": "趋势强度",
        "passed": trend["passed"],
        "detail": (
            "全部通过" if trend["passed"]
            else f"未通过: {', '.join(trend['failed_items'])}"
        ),
        "sub_checks": trend.get("checks", {}),
    })

    return gates


def _calc_avg_amplitude_diag(symbol: str, trade_date, local_data) -> float:
    """计算近5日平均振幅"""
    from datetime import date as dt_date, timedelta
    amps = []
    current = trade_date - timedelta(days=1)
    days_checked = 0
    while days_checked < 30 and len(amps) < 5:
        if current.weekday() < 5:
            try:
                bar = local_data.get_daily_quote(symbol, current)
                if bar and bar.get("high", 0) > 0 and bar.get("low", 0) > 0:
                    amp = (float(bar["high"]) - float(bar["low"])) / float(bar["low"])
                    amps.append(amp)
                days_checked += 1
            except Exception:
                days_checked += 1
        current -= timedelta(days=1)
    return sum(amps) / len(amps) if amps else 0.03


def _check_trend_strength_diag(symbol: str, trade_date, local_data) -> dict:
    """诊断用趋势强度检查（基于本地数据）"""
    from datetime import date as dt_date, timedelta

    checks = {}
    try:
        closes = _get_recent_closes(symbol, trade_date, None, local_data)

        # MA5 斜率
        if len(closes) >= 10:
            ma5_now = sum(closes[-5:]) / 5
            ma5_prev = sum(closes[-10:-5]) / 5
            if ma5_prev > 0:
                slope = (ma5_now - ma5_prev) / ma5_prev
                checks["ma5_slope"] = {
                    "passed": slope > 0,
                    "value": f"{slope:.2%}",
                    "threshold": "> 0",
                    "detail": f"MA5 {ma5_now:.2f} vs 前期{ma5_prev:.2f}",
                }
            else:
                checks["ma5_slope"] = {"passed": False, "value": "N/A", "threshold": "> 0", "detail": "MA5计算异常"}
        else:
            checks["ma5_slope"] = {"passed": False, "value": "N/A", "threshold": "> 0", "detail": f"数据不足(需≥10条,当前{len(closes)}条)"}

        # 量比
        try:
            vols = []
            current = trade_date
            days_collected = 0
            while days_collected < 30 and len(vols) < 6:
                if current.weekday() < 5:
                    bar = local_data.get_daily_quote(symbol, current)
                    if bar and bar.get("volume", 0) > 0:
                        vols.append(float(bar["volume"]))
                    days_collected += 1
                current -= timedelta(days=1)
            if len(vols) >= 6:
                vol_ratio = vols[0] / (sum(vols[1:6]) / 5) if sum(vols[1:6]) > 0 else 0
                checks["volume_ratio"] = {
                    "passed": vol_ratio > 0.8,
                    "value": f"{vol_ratio:.2f}",
                    "threshold": "> 0.8",
                }
            else:
                checks["volume_ratio"] = {"passed": True, "value": "N/A", "threshold": "> 0.8", "detail": "数据不足,跳过"}
        except Exception:
            checks["volume_ratio"] = {"passed": True, "value": "N/A", "threshold": "> 0.8", "detail": "计算跳过"}

        # 板块资金
        try:
            mapping = local_data.get_concept_mapping(symbol, trade_date)
            concepts = mapping.get("concepts", []) if mapping else []
            sector_flow = 0.0
            sector_name = ""
            if concepts:
                flows = local_data.get_concept_flow(trade_date, top_n=30) or []
                flow_map = {f.get("name", ""): f.get("main_net", 0) for f in flows}
                for cn in concepts[:3]:
                    cn_name = cn.get("concept_name", "") if isinstance(cn, dict) else str(cn)
                    sf = flow_map.get(cn_name, 0)
                    if sf != 0:
                        sector_flow = sf
                        sector_name = cn_name
                        break
            checks["sector_flow"] = {
                "passed": sector_flow > 0,
                "value": f"{sector_flow/1e8:.2f}亿" if sector_flow else "0",
                "threshold": "> 0",
                "detail": f"所属{sector_name}" if sector_name else "板块信息获取失败",
            }
        except Exception:
            checks["sector_flow"] = {"passed": True, "value": "N/A", "threshold": "> 0", "detail": "检查跳过"}

        # MA5 > MA20
        if len(closes) >= 20:
            ma5 = sum(closes[-5:]) / 5
            ma20 = sum(closes[-20:]) / 20
            checks["ma_align"] = {
                "passed": ma5 > ma20 and ma5 > 0,
                "value": f"MA5={ma5:.2f} MA20={ma20:.2f}",
                "threshold": "MA5 > MA20",
            }
        else:
            checks["ma_align"] = {"passed": False, "value": "N/A", "threshold": "MA5 > MA20", "detail": f"数据不足(需≥20条,当前{len(closes)}条)"}

        # 主力资金流向
        main_net_today, detail = _get_main_net_5d_diag(symbol, trade_date, None, None, local_data)
        checks["moneyflow"] = {
            "passed": main_net_today > 0,
            "value": f"{main_net_today/1e8:.2f}亿",
            "threshold": "> 0",
            "detail": "5日主力净流入",
        }

        # ── 综合判定：核心（MA5>MA20）+ 辅助 5选2 ──
        aux_keys = ['ma5_slope', 'volume_ratio', 'sector_flow', 'moneyflow']
        core_passed = checks.get('ma_align', {}).get('passed', False)
        aux_passed = sum(1 for k in aux_keys if checks.get(k, {}).get('passed', False))
        aux_total = len(aux_keys)

        if not core_passed:
            failed_items = ['ma_align(核心)'] + [k for k in aux_keys if not checks.get(k, {}).get('passed', False)]
        else:
            failed_items = [k for k, v in checks.items() if not v['passed']]

        all_passed = core_passed and aux_passed >= 2

        return {
            'passed': all_passed, 'failed_items': failed_items, 'checks': checks,
            'rule': 'core_plus_2aux', 'aux_passed': aux_passed, 'aux_total': aux_total,
            'core_passed': core_passed,
        }
    except Exception as e:
        return {"passed": False, "failed_items": ["exception"], "checks": {"error": str(e)}}


def _get_recent_closes(symbol: str, trade_date, cur_price: float, local_data) -> list:
    """获取最近25个交易日收盘价，含当日"""
    from datetime import date as dt_date, timedelta
    closes = []
    current = trade_date - timedelta(days=1)
    days_checked = 0
    while days_checked < 60 and len(closes) < 25:
        if current.weekday() < 5:
            try:
                bar = local_data.get_daily_quote(symbol, current)
                if bar and bar.get("close", 0) > 0:
                    closes.append(float(bar["close"]))
                days_checked += 1
            except Exception:
                days_checked += 1
        current -= timedelta(days=1)
    closes.reverse()
    if cur_price is not None and cur_price > 0:
        closes.append(cur_price)
    return closes


def _calc_rsi(closes: list, period: int = 6) -> float:
    """计算 RSI"""
    if len(closes) < period + 1:
        return 50.0
    gains = []
    losses = []
    for i in range(-period, 0):
        diff = closes[i] - closes[i-1]
        if diff > 0:
            gains.append(diff)
            losses.append(0)
        else:
            gains.append(0)
            losses.append(abs(diff))
    avg_gain = sum(gains) / period
    avg_loss = sum(losses) / period
    if avg_loss == 0:
        return 100.0
    rs = avg_gain / avg_loss
    return 100.0 - 100.0 / (1.0 + rs)


def _get_main_net_5d_diag(symbol: str, trade_date, hh, mm, local_data) -> tuple:
    """获取5日主力净流入（诊断用），返回 (net_amount, detail_str)"""
    from datetime import date as dt_date, timedelta
    try:
        main_net = 0.0
        days_count = 0
        cursor = trade_date
        while days_count < 5:
            if cursor.weekday() < 5:
                mf = local_data.get_moneyflow(symbol, cursor)
                if mf:
                    lg = float(mf.get("buy_elg_amount", 0) or 0) - float(mf.get("sell_elg_amount", 0) or 0)
                    md = float(mf.get("buy_lg_amount", 0) or 0) - float(mf.get("sell_lg_amount", 0) or 0)
                    day_val = lg + md
                    if cursor == trade_date and hh is not None and mm is not None:
                        try:
                            wi = local_data.get_moneyflow_intraday_weight(symbol, trade_date, hh, mm)
                            if wi:
                                day_val *= wi["weight"]
                        except Exception:
                            pass
                    main_net += day_val
                    days_count += 1
            cursor -= timedelta(days=1)
        return main_net, f"累计{days_count}日"
    except Exception:
        return 0.0, "计算异常"

